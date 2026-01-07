import re
import os
import json
import io
import random
import time
import subprocess
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.alert import Alert
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from flask import Flask, render_template, request, jsonify, Response
import uuid
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === Configuration ===
GOOGLE_API_KEYS = [
    "AIzaSyDASUJ9-Q1kw0uYoUYuIpNZmBBvG-0PlCE",
    "AIzaSyAeHrqRKZ1nYn_nNN8KXrgDrhX8_hy-bKo",
    "AIzaSyCTWb-yJEKMc6ff9CXiW-jEWol05w7VldU",
    "AIzaSyB8FHdHOcHygkkFxOitFdBxuT9MMwLwqoQ"
]
SEARCH_ENGINE_ID = "a6cea8f5219ce4ccb"
MAX_DEPTH = 3
CRAWL_DELAY = 0.5
RETRY_ATTEMPTS = 2
TIMEOUT = 30
DEBUG_DIR = "debug_html"
MODEL_NAME = "meta-llama/Meta-Llama-3.1-8B-Instruct"
MAX_SELENIUM_WORKERS = 4
BATCH_SIZE = 5

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:120.0) Gecko/20100101 Firefox/120.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Edge/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Mobile/15E148 Safari/604.1",
    "Mozilla/5.0 (iPad; CPU OS 17_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Mobile/15E148 Safari/604.1"
]


# Compiled regex patterns
EMAIL_REGEX = re.compile(
    r'(?:'  # Start non-capturing group
    r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,6}\b|'  # Standard email
    r'[a-zA-Z0-9._%+-]+\s*$$at$$\s*[a-zA-Z0-9.-]+\s*$$dot$$\s*[a-zA-Z]{2,6}\b|'  # (at)/(dot)
    r'[a-zA-Z0-9._%+-]+%40[a-zA-Z0-9.-]+\.[a-zA-Z]{2,6}\b|'  # Encoded @ (%40)
    r'"email":"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,6}"'  # JSON email
    r')',  # Close non-capturing group
    re.IGNORECASE)

PHONE_REGEX = re.compile(
    r'(?<![\w\-/])'  # Negative lookbehind to avoid matching within words or paths
    r'(?:\+(?:1|44|61|91|977|86|81|49|33|39|34|7|55|52|54|27|20|234|254|256|233|63|92|94|95|880|60|65|66|84|852|853|886|82|357|30|31|32|41|43|45|46|47|48|351|420|421|36|695|386|383|389|355|377|376|378|379|380|381|382|385|387|359|40|373))'  # Country codes
    r'[-.\s]?\d{1,4}[-.\s]?\d{2,4}[-.\s]?\d{2,4}(?:[-.\s]?\d{2,4})?(?:\s*(?:ext\.?|extension|x)\s*\d{1,5})?'  # Phone number pattern
    r'|tel:\+\d{1,15}'  # tel: links with country code
    r'|"(?:phone|tel)":\s*"\+[^"]*"'  # JSON phone with country code
    r'(?!\w)',  # Negative lookahead to avoid trailing characters
    re.IGNORECASE
)

ADDRESS_REGEX = re.compile(
    r'\b(?:\d{1,5}\s+)?[\w\s.-]+(?:\s+(?:St|Ave|Rd|Blvd|Ln|Dr|Ct|Pl|Way|Ter|Cir|Pkwy|Sq)(?:\.|(?:reet|venue|oad|oulevard|ane|rive|ourt|lace|errace|ircle|arkway|quare))?)?(?:[\s,]+[\w\s.-]+){1,4}(?:\s+\d{5}(?:-\d{4})?)?\b',
    re.IGNORECASE)

SOCIAL_REGEXES = {
    "facebook": re.compile(
        r'(?:https?://)?(?:[a-z]{2,3}\.)?facebook\.com/(?:profile\.php\?id=\d+|(?:pages/)?[A-Za-z0-9._-]+/?)(?:\?[^"\']*)?$',
        re.IGNORECASE
    ),
    "instagram": re.compile(
        r'(?:https?://)?(?:[a-z]{2,3}\.)?instagram\.com/[A-Za-z0-9._-]+/?(?:\?.*)?$',
        re.IGNORECASE
    ),
    "linkedin": re.compile(
        r'(?:https?://)?(?:[a-z]{2,3}\.)?linkedin\.com/(?:company|in)/[A-Za-z0-9._-]+/?(?:\?.*)?$',
        re.IGNORECASE
    ),
    "youtube": re.compile(
        r'(?:https?://)?(?:[a-z]{2,3}\.)?youtube\.com/(?:channel|user|c)/[A-Za-z0-9._-]+/?(?:\?.*)?$',
        re.IGNORECASE
    ),
    "pinterest": re.compile(
        r'(?:https?://)?(?:[a-z]{2,3}\.)?pinterest\.com/[A-Za-z0-9._-]+/?(?:\?.*)?$',
        re.IGNORECASE
    ),
}
LINK_REGEX = re.compile(r'(?:href|url|Link|to)\s*[:=]\s*[\'"]([^"\']+)[\'"]', re.IGNORECASE)

REACT_ATTRIBUTES = [
    "id", "class", "href", "src", "for", "type", "name", "value", "placeholder",
    "role", "aria-label", "aria-hidden", "aria-describedby", "data-testid", "data-component",
    "product", "backgroundColor", "themeColor", "titleColor", "textColor", "footerURL"
]

# Ensure debug directory exists
if not os.path.exists(DEBUG_DIR):
    os.makedirs(DEBUG_DIR)

# Initialize Flask app
app = Flask(__name__)
app.secret_key = str(uuid.uuid4())

# Store results globally for export
current_results = []

# Initialize Llama Model Pipeline
try:
    import transformers
    import torch

    pipeline = transformers.pipeline(
        "text-generation",
        model=MODEL_NAME,
        model_kwargs={"torch_dtype": torch.bfloat16},
        device_map="auto",
    )
except Exception as e:
    print(f"❌ Failed to load Llama model: {e}")
    pipeline = None
    
def fallback_keyword_matching(links):
    """Fallback to simple keyword matching when Llama fails"""
    about_links = []
    contact_links = []
    
    about_patterns = [
        r'.*about.*', r'.*story.*', r'.*team.*', r'.*company.*', r'.*mission.*',
        r'.*vision.*', r'.*who.*we.*are.*', r'.*history.*', r'.*leadership.*',
        r'.*values.*', r'.*culture.*', r'.*philosophy.*'
    ]
    
    contact_patterns = [
        r'.*contact.*', r'.*get.*in.*touch.*', r'.*reach.*us.*', r'.*support.*',
        r'.*help.*', r'.*customer.*service.*', r'.*inquiries.*', r'.*feedback.*'
    ]
    
    for link in links:
        url = link.get('url', '').lower()
        text = link.get('text', '').lower()
        
        # Check for about pages
        for pattern in about_patterns:
            if re.search(pattern, url) or re.search(pattern, text):
                about_links.append(link)
                break
        
        # Check for contact pages
        for pattern in contact_patterns:
            if re.search(pattern, url) or re.search(pattern, text):
                contact_links.append(link)
                break
    
    return {
        "about": about_links,
        "contact": contact_links
    }

# Load country codes for phone validation
COUNTRY_CODES = {}
def load_country_codes():
    """Load country codes from countryCode.txt file"""
    global COUNTRY_CODES
    try:
        country_code_file = os.path.join(os.path.dirname(__file__), "countryCode.txt")
        with open(country_code_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if '\t' in line:
                    parts = line.split('\t')
                    if len(parts) == 2:
                        country = parts[0].strip()
                        codes = parts[1].strip()
                        # Handle multiple codes (e.g., "1-787, 1-939")
                        for code in codes.split(','):
                            code = code.strip()
                            COUNTRY_CODES[code] = country
        custom_print(f"✅ Loaded {len(COUNTRY_CODES)} country codes")
    except Exception as e:
        print(f"❌ Failed to load country codes: {e}")

# Validate phone number with country codes
def validate_phone_with_country_code(phone):
    """Validate phone number against country codes and return phone with country info"""
    # Extract potential country code from phone number
    phone_clean = re.sub(r'[^\d+]', '', phone)

    # Check if phone starts with +
    if phone_clean.startswith('+'):
        phone_clean = phone_clean[1:]

    # Try to match country codes (from longest to shortest)
    matched_country = None
    matched_code = None

    # Sort codes by length (longest first) to match more specific codes first
    sorted_codes = sorted(COUNTRY_CODES.keys(), key=lambda x: len(x.replace('-', '')), reverse=True)

    for code in sorted_codes:
        code_clean = code.replace('-', '')
        if phone_clean.startswith(code_clean):
            matched_code = code
            matched_country = COUNTRY_CODES[code]
            break

    if matched_country:
        return {
            "phone": phone,
            "country_code": matched_code,
            "country": matched_country,
            "is_valid": True
        }
    else:
        return {
            "phone": phone,
            "country_code": "Unknown",
            "country": "Unknown",
            "is_valid": False
        }

# Identify About/Contact links using Llama
def identify_about_contact_links(links):
    """
    Use Llama to identify which links are About Us or Contact Us pages
    Args:
        links: List of dicts with 'url' and 'text' keys
    Returns:
        Dict with 'about' and 'contact' lists containing identified links
    """
    if not pipeline or not links:
        custom_print("❌ Llama model not loaded or no links provided")
        return {"about": [], "contact": []}

    # Pre-filtering logic
    about_keywords = [
        'about', 'story', 'team', 'company', 'mission', 'vision', 'who-we-are',
        'our-story', 'our-team', 'our-mission', 'history', 'leadership', 'values',
        'meet-the-team', 'meet-us', 'who-are-we', 'culture', 'philosophy'
    ]

    contact_keywords = [
        'contact', 'get-in-touch', 'reach-us', 'support', 'help', 'customer-service',
        'contact-us', 'reach-out', 'talk-to-us', 'customer-support', 'assistance',
        'inquiries', 'feedback', 'get-help', 'support-center'
    ]

    # Keywords to exclude
    exclude_keywords = [
        'product', 'cart', 'checkout', 'login', 'register', 'sign-in', 'sign-up',
        'account', 'wishlist', 'shop', 'collection', 'category', 'search', 'filter',
        'sort', 'view', 'add-to-cart', 'buy', 'purchase', 'price', 'sale', 'deal',
        'shipping', 'return', 'refund', 'tracking', 'order', '/products/', '/cart',
        '/checkout', '/account', '/login', '/register', '/search', 'color', 'size',
        'material', 'fabric', 'quick-ship', 'ready-to-ship', 'gift-card', 'blog',
        'news', 'article', 'post', 'kitchen', 'dining', 'living', 'bedroom', 'bathroom',
        'entryway', 'hallway', 'runner', 'rug', 'loom', 'wool', 'natural', 'red',
        'orange', 'yellow', 'green', 'blue', 'purple', 'pink', 'black', 'brown',
        'grey', 'tan', 'minis', '4x6', '5x7', '6x9', '8x8', '8x10', '9x12', '10x10'
    ]

    filtered_links = []
    filtered_indices = []  # Keep track of original indices
    
    for i, link in enumerate(links):
        url = link.get('url', '').lower()
        text = link.get('text', '').lower()

        # Check if should be excluded
        should_exclude = any(keyword in url or keyword in text for keyword in exclude_keywords)
        if should_exclude:
            continue

        # Check if matches about/contact keywords
        is_about = any(keyword in url or keyword in text for keyword in about_keywords)
        is_contact = any(keyword in url or keyword in text for keyword in contact_keywords)

        if is_about or is_contact:
            filtered_links.append(link)
            filtered_indices.append(i)  # Store the original index

    custom_print(f"🔍 Pre-filtered from {len(links)} to {len(filtered_links)} potentially relevant links")

    if not filtered_links:
        custom_print("⚠️ No links matched pre-filtering criteria")
        return {"about": [], "contact": []}

    # Prepare link information for Llama
    links_text = []
    for i, link in enumerate(filtered_links):
        url = link.get('url', '')
        text = link.get('text', '')
        links_text.append(f"{i}. URL: {url} | Link Text: {text}")

    links_input = "\n".join(links_text[:50])  # Limit to 50 links

    prompt = f"""From the following navigation/footer links, identify which pages are "About Us" or "Contact Us" pages.

About Us pages include: company information, team, story, mission, values, history, who we are, etc.
Contact Us pages include: contact forms, customer support, get in touch, reach us, help, inquiries, etc.

Look at both the URL and link text to decide.

Pre-filtered Links:
{links_input}

Return a JSON object with two keys: "about" and "contact". Each key should have an array of the indices (numbers) from the list above that correspond to About or Contact pages.

Example response format:
{{
  "about": [0, 2, 5],
  "contact": [1, 3]
}}

Return ONLY the JSON object, no additional text or explanations."""

    messages = [
        {"role": "system", "content": "You are a helpful assistant that analyzes website navigation links to identify About and Contact pages. Return ONLY valid JSON."},
        {"role": "user", "content": prompt}
    ]

    try:
        outputs = pipeline(
            messages,
            max_new_tokens=512,
            temperature=0.1,  # Lower temperature for more consistent output
        )
        content = outputs[0]["generated_text"][-1]["content"]
        
        # Log the raw output for debugging
        custom_print(f"🔗 Llama raw output: {content}")
        
        # Save to debug file
        with open(os.path.join(DEBUG_DIR, "llama_link_identification.txt"), "a", encoding="utf-8") as f:
            f.write(f"Input:\n{links_input}\n\nLlama output:\n{content}\n\n{'='*80}\n\n")

        # Try to extract JSON from the output
        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        if json_match:
            json_str = json_match.group(0)
            try:
                result = json.loads(json_str)
                about_indices = result.get("about", [])
                contact_indices = result.get("contact", [])
                
                custom_print(f"📊 Parsed: {len(about_indices)} About indices, {len(contact_indices)} Contact indices")
                
                # Map filtered indices back to original links
                identified = {
                    "about": [],
                    "contact": []
                }
                
                for idx in about_indices:
                    if isinstance(idx, int) and idx < len(filtered_links):
                        # Get the original link using the stored original index
                        original_idx = filtered_indices[idx]
                        if original_idx < len(links):
                            identified["about"].append(links[original_idx])
                
                for idx in contact_indices:
                    if isinstance(idx, int) and idx < len(filtered_links):
                        original_idx = filtered_indices[idx]
                        if original_idx < len(links):
                            identified["contact"].append(links[original_idx])
                
                custom_print(f"✅ Identified {len(identified['about'])} About pages and {len(identified['contact'])} Contact pages")
                return identified
                
            except json.JSONDecodeError as e:
                custom_print(f"❌ Failed to parse JSON: {e}")
                custom_print(f"📄 Content was: {json_str}")
        
        # Fallback if no JSON found
        custom_print("⚠️ No valid JSON found in model output, falling back to keyword matching")
        return fallback_keyword_matching(filtered_links)

    except Exception as e:
        custom_print(f"❌ Error identifying links with Llama: {e}")
        # Fallback to keyword matching
        return fallback_keyword_matching(filtered_links)

# Custom print function to capture console output
process_log = []

def validate_social_link(url, platform):
    """Validate and clean social media links to ensure they're profile/page URLs, not media URLs"""
    
    # Skip if URL contains media file extensions or CDN patterns
    media_patterns = [
        r'\.(?:jpg|jpeg|png|gif|mp4|webp|svg)(?:\?|$)',  # Media files
        r'scontent\.cdninstagram\.com',  # Instagram CDN
        r'scontent\..*\.fna\.fbcdn\.net',  # Facebook CDN
        r'fbcdn\.net',  # Facebook CDN
        r'/v/t\d+\.',  # Instagram media paths
        r'/o1/v/t\d+/',  # Instagram media paths
        r'/reel[s]?/',  # Instagram reels
        r'/p/',  # Instagram posts
        r'/stories/',  # Instagram stories
        r'/explore/',  # Instagram explore
        r'/direct/',  # Instagram direct messages
    ]
    
    for pattern in media_patterns:
        if re.search(pattern, url, re.IGNORECASE):
            custom_print(f"🚫 Rejected {platform} URL due to media pattern match: {url} (pattern: {pattern})")
            return None
    
    # Platform-specific validation
    if platform == "instagram":
        instagram_profile_pattern = r'^(?:https?://)?(?:[a-z]{2,3}\.)?instagram\.com/([A-Za-z0-9._-]+)/?(?:\?.*)?$'
        match = re.match(instagram_profile_pattern, url, re.IGNORECASE)
        if match:
            username = match.group(1)
            # Skip generic/system usernames and paths
            skip_usernames = ['tv', 'reel', 'reels', 'stories', 'explore', 'accounts', 'direct', 'developer']
            if username.lower() not in skip_usernames and len(username) > 1:
                custom_print(f"✅ Validated Instagram URL: {url} -> Username: {username}")
                return f"https://www.instagram.com/{username}/"
            else:
                custom_print(f"🚫 Rejected Instagram URL due to invalid username: {url} (username: {username})")
                return None
        custom_print(f"🚫 Rejected Instagram URL due to pattern mismatch: {url}")
        return None
    
    elif platform == "facebook":
        facebook_pattern = r'^(?:https?://)?(?:[a-z]{2,3}\.)?facebook\.com/(?:profile\.php\?id=(\d+)(?:&.*)?$|(?:pages/)?([A-Za-z0-9._-]+)/?(?:\?.*)?$)'
        match = re.match(facebook_pattern, url, re.IGNORECASE)
        if match:
            if match.group(1):  # ID-based profile
                custom_print(f"✅ Validated Facebook URL (ID-based): {url}")
                return f"https://www.facebook.com/profile.php?id={match.group(1)}"
            elif match.group(2):  # Username-based profile or page
                custom_print(f"✅ Validated Facebook URL (username-based): {url}")
                return f"https://www.facebook.com/{match.group(2)}/"
        custom_print(f"🚫 Rejected Facebook URL due to pattern mismatch: {url}")
        return None
    
    elif platform == "pinterest":
        pinterest_pattern = r'^(?:https?://)?(?:[a-z]{2,3}\.)?pinterest\.com/([A-Za-z0-9._-]+)/?(?:\?.*)?$'
        match = re.match(pinterest_pattern, url, re.IGNORECASE)
        if match:
            username = match.group(1)
            custom_print(f"✅ Validated Pinterest URL: {url} -> Username: {username}")
            return f"https://www.pinterest.com/{username}/"
        custom_print(f"🚫 Rejected Pinterest URL due to pattern mismatch: {url}")
        return None
    
    elif platform == "linkedin":
        linkedin_pattern = r'^(?:https?://)?(?:[a-z]{2,3}\.)?linkedin\.com/(?:company|in)/([A-Za-z0-9._-]+)/?(?:\?.*)?$'
        match = re.match(linkedin_pattern, url, re.IGNORECASE)
        if match:
            path = match.group(1)
            custom_print(f"✅ Validated LinkedIn URL: {url} -> Path: {path}")
            return f"https://www.linkedin.com/{path}/"
        custom_print(f"🚫 Rejected LinkedIn URL due to pattern mismatch: {url}")
        return None
    
    elif platform == "youtube":
        youtube_pattern = r'^(?:https?://)?(?:[a-z]{2,3}\.)?youtube\.com/(?:channel|user|c)/([A-Za-z0-9._-]+)/?(?:\?.*)?$'
        match = re.match(youtube_pattern, url, re.IGNORECASE)
        if match:
            path = match.group(1)
            custom_print(f"✅ Validated YouTube URL: {url} -> Path: {path}")
            return f"https://www.youtube.com/{path}/"
        custom_print(f"🚫 Rejected YouTube URL due to pattern mismatch: {url}")
        return None
    
    # For other platforms, return as-is if it matches the platform's regex
    if SOCIAL_REGEXES.get(platform, re.compile(r'')).match(url):
        custom_print(f"✅ Validated {platform} URL (fallback): {url}")
        return url
    custom_print(f"🚫 Rejected {platform} URL due to regex mismatch: {url}")
    return None

def custom_print(*args, **kwargs):
    message = " ".join(str(arg) for arg in args)
    process_log.append({"timestamp": time.strftime("%Y-%m-%d %H:%M:%S %Z"), "message": message})
    print(*args, **kwargs)

# Load country codes at startup
load_country_codes()

def get_all_google_domains(query, company_name=None):
    """Get all domains from Google search results and validate them"""
    social_media_domains = [
        'facebook.com', 'instagram.com', 'linkedin.com', 'youtube.com', 'pinterest.com',
        'twitter.com', 'tiktok.com', 'snapchat.com', 'reddit.com'
    ]
    
    company_clean = ""
    company_words = []
    if company_name:
        company_clean = re.sub(r'\s*(pvt|ltd|limited|private|inc|incorporated|co|company|corp|corporation|llc|gmbh|plc)\s*\.?$', 
                              '', company_name, flags=re.IGNORECASE)
        company_clean = company_clean.lower().strip()
        company_words = re.findall(r'\b\w+\b', company_clean)
    
    custom_print(f"🔍 Searching Google for all domains matching: '{company_clean}' (words: {company_words})")
    
    all_domains = []
    
    for api_key in GOOGLE_API_KEYS:
        try:
            service = build("customsearch", "v1", developerKey=api_key)
            res = service.cse().list(q=query, cx=SEARCH_ENGINE_ID, num=10).execute()
            items = res.get("items", [])
            
            for item in items:
                link = item.get("link", "")
                title = item.get("title", "").lower()
                snippet = item.get("snippet", "").lower()
                parsed = urlparse(link)
                domain_name = parsed.netloc.lower()
                domain = f"{parsed.scheme}://{parsed.netloc}/"
                
                # Skip social media domains
                if any(social_domain in domain_name for social_domain in social_media_domains):
                    continue
                
                # Extract domain without TLD and www for matching
                domain_base = re.sub(r'^www\.', '', domain_name)
                domain_base = re.sub(r'\.(com|org|net|io|co|in|us|uk|au|ca|de|fr|jp|biz|info|mobi|name|tv|cc|ws)$', '', domain_base)
                
                # Check if this domain matches the company
                matches_company = False
                if company_words:
                    matches = []
                    for word in company_words:
                        if word in domain_base:
                            matches.append(word)
                    
                    # For multi-word companies, require ALL words to match
                    if len(company_words) >= 2:
                        matches_company = (len(matches) == len(company_words))
                    else:
                        matches_company = (len(matches) == 1)
                
                domain_info = {
                    'domain': domain,
                    'domain_name': domain_name,
                    'domain_base': domain_base,
                    'title': title[:100],
                    'matches_company': matches_company,
                    'full_match': False
                }
                
                # Mark as full match if ALL company words are in domain
                if company_words and matches_company:
                    domain_info['full_match'] = True
                    custom_print(f"  ✅ FULL MATCH: {domain_name}")
                else:
                    custom_print(f"  ⚠️ Partial/No match: {domain_name}")
                
                all_domains.append(domain_info)
            
            break  # Successfully got results, stop trying API keys
            
        except HttpError as e:
            if e.resp.status == 429:
                custom_print(f"❌ Google API error with key {api_key}: Quota exceeded. Trying next API key...")
                continue
            else:
                custom_print(f"❌ Google API error with key {api_key}: {e}")
                return []
        except Exception as e:
            custom_print(f"❌ Google API error with key {api_key}: {e}")
            return []
    
    return all_domains

def find_domain_google(query, company_name=None):
    """Find domain using Google Search API with multiple search approaches"""
    
    # Clean company name for matching
    company_clean = ""
    company_words = []
    if company_name:
        company_clean = re.sub(r'\s*(pvt|ltd|limited|private|inc|incorporated|co|company|corp|corporation|llc|gmbh|plc)\s*\.?$', 
                              '', company_name, flags=re.IGNORECASE)
        company_clean = company_clean.lower().strip()
        company_words = re.findall(r'\b\w+\b', company_clean)
    
    custom_print(f"🔍 Searching for domain matching: '{company_clean}' (words: {company_words})")
    
    # Try multiple search queries
    search_queries = [
        query,  # Original query
        f"{company_name} website",  # Add "website" keyword
        f"{company_name} official website",  # Add "official website"
        f"{company_clean.replace(' ', '')}.com",  # Try domain pattern
        f"{company_clean} com",  # Try with "com"
    ]
    
    all_domains = []
    
    for search_query in search_queries:
        custom_print(f"  Trying search query: '{search_query}'")
        
        for api_key in GOOGLE_API_KEYS:
            try:
                service = build("customsearch", "v1", developerKey=api_key)
                res = service.cse().list(q=search_query, cx=SEARCH_ENGINE_ID, num=10).execute()
                items = res.get("items", [])
                
                for item in items:
                    link = item.get("link", "")
                    title = item.get("title", "").lower()
                    snippet = item.get("snippet", "").lower()
                    parsed = urlparse(link)
                    domain_name = parsed.netloc.lower()
                    
                    # Skip social media and common unrelated sites
                    social_media = ['facebook', 'instagram', 'linkedin', 'youtube', 'pinterest', 
                                   'twitter', 'tiktok', 'snapchat', 'reddit', 'wikipedia']
                    if any(social in domain_name for social in social_media):
                        continue
                    
                    domain = f"{parsed.scheme}://{parsed.netloc}/"
                    
                    # Extract domain without TLD and www for matching
                    domain_base = re.sub(r'^www\.', '', domain_name)
                    domain_base = re.sub(r'\.(com|org|net|io|co|in|us|uk|au|ca|de|fr|jp|biz|info|mobi|name|tv|cc|ws)$', '', domain_base)
                    
                    # Check if this domain matches the company
                    matches_company = False
                    match_score = 0
                    
                    if company_words:
                        matches = []
                        for word in company_words:
                            if len(word) > 3 and word in domain_base:
                                matches.append(word)
                                match_score += 10
                        
                        # Also check for combined words
                        company_combined = company_clean.replace(" ", "")
                        if company_combined in domain_base:
                            match_score += 50  # Big bonus for exact combined match
                            matches_company = True
                        elif len(matches) >= len(company_words):
                            match_score += 30  # All words match
                            matches_company = True
                        elif len(matches) > 0:
                            match_score += 15  # Some words match
                    
                    # Check title and snippet for company mention
                    for word in company_words:
                        if word in title:
                            match_score += 5
                        if word in snippet:
                            match_score += 3
                    
                    # Skip low scoring domains
                    if match_score < 10:
                        continue
                    
                    domain_info = {
                        'domain': domain,
                        'domain_name': domain_name,
                        'title': title[:100],
                        'match_score': match_score,
                        'matches_company': matches_company,
                        'search_query': search_query
                    }
                    
                    # Check if we already have this domain
                    if not any(d['domain'] == domain for d in all_domains):
                        all_domains.append(domain_info)
                        custom_print(f"    Found: {domain_name} (score: {match_score})")
                
                break  # Successfully got results with this API key
                
            except HttpError as e:
                if e.resp.status == 429:
                    custom_print(f"❌ Google API error with key {api_key}: Quota exceeded. Trying next API key...")
                    continue
                else:
                    custom_print(f"❌ Google API error with key {api_key}: {e}")
                    continue
            except Exception as e:
                custom_print(f"❌ Google API error with key {api_key}: {e}")
                continue
    
    # Sort domains by match score
    all_domains.sort(key=lambda x: x['match_score'], reverse=True)
    
    custom_print(f"\n📊 Search Results Analysis:")
    custom_print(f"  Total domains found: {len(all_domains)}")
    
    if all_domains:
        custom_print(f"\n✅ Top matching domains:")
        for i, domain_info in enumerate(all_domains[:5], 1):
            custom_print(f"  {i}. {domain_info['domain']} (score: {domain_info['match_score']})")
            custom_print(f"     Title: {domain_info['title']}")
        
        # Return the best matching domain
        best = all_domains[0]
        if best['match_score'] >= 20:  # Reasonable threshold
            custom_print(f"\n✅ Using best matching domain: {best['domain']} (score: {best['match_score']})")
            return best['domain']
        else:
            custom_print(f"\n⚠️ Best domain score too low: {best['match_score']} (threshold: 20)")
            custom_print(f"   Best available: {best['domain']}")
            return None
    
    custom_print("❌ No suitable domains found in any search")
    
    # Fallback: Try direct domain construction
    if company_words and len(company_words) >= 2:
        possible_domains = [
            f"https://{company_clean.replace(' ', '').lower()}.com",
            f"https://{company_clean.replace(' ', '-').lower()}.com",
            f"https://www.{company_clean.replace(' ', '').lower()}.com",
            f"https://www.{company_clean.replace(' ', '-').lower()}.com",
        ]
        
        custom_print(f"\n🔄 Trying direct domain construction...")
        for test_domain in possible_domains:
            custom_print(f"  Testing: {test_domain}")
            if is_link_accessible(test_domain, max_retries=1):
                custom_print(f"✅ Direct domain accessible: {test_domain}")
                return test_domain
    
    return None

# Initialize Selenium driver with optimized configuration
def init_selenium_driver():
    options = Options()
    
    # Basic headless options
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")  # More realistic window size
    options.add_argument("--disable-dev-shm-usage")
    
    # Anti-detection measures
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-extensions-except")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-plugins-discovery")
    options.add_argument("--disable-web-security")
    options.add_argument("--allow-running-insecure-content")
    options.add_argument("--disable-features=VizDisplayCompositor")
    options.add_argument("--disable-background-networking")
    options.add_argument("--disable-client-side-phishing-detection")
    options.add_argument("--disable-hang-monitor")
    options.add_argument("--disable-prompt-on-repost")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    
    # Random User-Agent
    user_agent = random.choice(USER_AGENTS)
    options.add_argument(f"--user-agent={user_agent}")
    
    # Additional headers to appear more legitimate
    options.add_argument("--accept-language=en-US,en;q=0.9")
    options.add_argument("--accept-encoding=gzip, deflate, br")
    
    # Disable automation indicators
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    
    # Enable images and JavaScript for more realistic browsing (removed disable arguments)
    options.add_argument("--enable-javascript")
    
    # Performance optimizations while keeping functionality
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-renderer-backgrounding")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--log-level=3")
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    
    # Execute script to hide webdriver property
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
    # Set realistic timeouts
    driver.set_page_load_timeout(45)  # Increased timeout
    driver.implicitly_wait(10)
    
    return driver

def access_with_referer(driver, url):
    """Access URL with a fake referer"""
    from urllib.parse import urlparse
    parsed_url = urlparse(url)
    base_domain = f"{parsed_url.scheme}://{parsed_url.netloc}"
    
    try:
        # Set referer header via CDP
        driver.execute_cdp_cmd('Network.setUserAgentOverride', {
            "userAgent": random.choice(USER_AGENTS)
        })
    except:
        pass
    
    # First visit the base domain
    try:
        driver.get(base_domain)
        time.sleep(2)
    except:
        pass
    
    # Then visit the target URL
    driver.get(url)

def staged_access(driver, url):
    """Access URL in stages, mimicking human behavior"""
    from urllib.parse import urlparse
    parsed_url = urlparse(url)
    base_domain = f"{parsed_url.scheme}://{parsed_url.netloc}"
    
    stages = [
        base_domain,
        f"{base_domain}/",
        url
    ]
    
    for stage in stages:
        try:
            driver.get(stage)
            time.sleep(random.uniform(1, 3))
            
            # Simulate some human-like actions
            driver.execute_script("window.scrollTo(0, Math.floor(Math.random() * 200));")
            time.sleep(random.uniform(0.5, 1.5))
        except:
            continue

def access_with_mobile_ua(driver, url):
    """Access URL with mobile user agent"""
    mobile_ua = "Mozilla/5.0 (iPhone; CPU iPhone OS 17_1 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.1 Mobile/15E148 Safari/604.1"
    
    try:
        driver.execute_cdp_cmd('Network.setUserAgentOverride', {
            "userAgent": mobile_ua
        })
    except:
        pass
    
    driver.get(url)

# Render page with Selenium
def get_rendered_html(url, driver, retries=RETRY_ATTEMPTS):
    debug_data = {
        "url": url,
        "dynamic_html": "",
        "links": [],
        "scripts": [],
        "raw_socials": [],
        "raw_emails": [],
        "raw_phones": [],
        "raw_addresses": [],
        "shadow_content": [],
        "alerts": [],
        "consolidated_links": [],
        "iframes": [],
        "access_methods_tried": [],
        "error_details": []  # New key to store detailed error information
    }
    
    consolidated_links = set()
    
    # Try different approaches for 403 errors
    access_methods = [
        ('direct', lambda: driver.get(url)),
        ('with_referer', lambda: access_with_referer(driver, url)),
        ('staged_access', lambda: staged_access(driver, url)),
        ('mobile_user_agent', lambda: access_with_mobile_ua(driver, url))
    ]
    
    for method_name, access_method in access_methods:
        for attempt in range(retries):
            try:
                debug_data["access_methods_tried"].append(f"{method_name}_attempt_{attempt + 1}")
                custom_print(f"🔄 Trying {method_name} method for {url} (attempt {attempt + 1})")
                
                # Execute the access method
                access_method()
                
                # Add random human-like delay
                time.sleep(random.uniform(2, 5))
                
                # Check if we got content
                current_url = driver.current_url
                page_source = driver.page_source
                
                # Check for common error indicators
                error_indicators = ['403 forbidden', '403 error', 'access denied', 'unauthorized', 'blocked']
                if any(indicator in page_source.lower() for indicator in error_indicators):
                    # Capture HTTP status if possible
                    status_code = None
                    try:
                        response = requests.head(url, timeout=10, allow_redirects=True)
                        status_code = response.status_code
                    except Exception as e:
                        custom_print(f"❌ Failed to get HTTP status for {url}: {e}")
                    
                    error_detail = {
                        "method": method_name,
                        "attempt": attempt + 1,
                        "error_type": "Error Page",
                        "message": f"Received error page with indicators: {', '.join(indicator for indicator in error_indicators if indicator in page_source.lower())}",
                        "http_status": status_code if status_code else "Unknown",
                        "current_url": current_url
                    }
                    debug_data["error_details"].append(error_detail)
                    custom_print(f"❌ {method_name} method got error page for {url}: {error_detail['message']} (HTTP Status: {error_detail['http_status']})")
                    continue
                
                if len(page_source) < 500:
                    error_detail = {
                        "method": method_name,
                        "attempt": attempt + 1,
                        "error_type": "Minimal Content",
                        "message": f"Received minimal content (length: {len(page_source)} bytes)",
                        "http_status": None,
                        "current_url": current_url
                    }
                    debug_data["error_details"].append(error_detail)
                    custom_print(f"❌ {method_name} method got minimal content for {url}: {error_detail['message']}")
                    continue
                
                custom_print(f"✅ {method_name} method successful for {url}")
                
                # Wait for page elements (unchanged)
                try:
                    WebDriverWait(driver, TIMEOUT).until(
                        EC.any_of(
                            EC.presence_of_element_located((By.TAG_NAME, "body")),
                            EC.presence_of_element_located((By.CSS_SELECTOR, (
                                "footer, [class*='footer' i], [id*='footer' i], "
                                "[class*='Footer' i], [id*='Footer' i], "
                                "explorug-footer, explorug-contact, #root explorug-footer, "
                                "#root explorug-contact, #root [class*='footer' i], #root [class*='Footer' i], "
                                "[class*='bottom' i], [class*='foot' i], #root > *:last-child"
                            )))
                        )
                    )
                except Exception as e:
                    debug_data["alerts"].append(f"Timeout waiting for page elements: {str(e)}")
                
                # Handle alerts (unchanged)
                try:
                    alert = Alert(driver)
                    alert_text = alert.text
                    debug_data["alerts"].append(f"Alert on {method_name}: {alert_text}")
                    custom_print(f"⚠️ Alert found on {url}: {alert_text}")
                    alert.dismiss()
                except:
                    pass
                
                # Trigger dynamic content loading and interactions (unchanged)
                driver.execute_script("""
                    window.scrollTo(0, document.body.scrollHeight);
                    document.querySelectorAll('button, [role="button"], [class*="load-more"], [class*="show-more"], [class*="footer" i], [class*="Footer" i], explorug-footer, explorug-contact').forEach(btn => {
                        try { btn.click(); } catch(e) {}
                    });
                    // Trigger custom events for frameworks like React
                    ['click', 'mouseover'].forEach(eventType => {
                        document.querySelectorAll('#root, footer, explorug-footer, explorug-contact, [class*="footer" i], [class*="Footer" i]').forEach(elem => {
                            try {
                                let event = new Event(eventType, { bubbles: true });
                                elem.dispatchEvent(event);
                            } catch(e) {}
                        });
                    });
                """)
                time.sleep(2)
                
                # Shadow DOM and iframe processing (unchanged)
                shadow_js_script = """
                    function getShadowHTML(host) {
                        if (!host.shadowRoot) return null;
                        const result = [];
                        try {
                            [...host.shadowRoot.children].forEach(child => {
                                result.push(child.outerHTML);
                                child.querySelectorAll('*').forEach(nested => {
                                    if (nested.shadowRoot) {
                                        const nestedShadow = getShadowHTML(nested);
                                        if (nestedShadow) result.push(...nestedShadow);
                                    }
                                });
                            });
                        } catch(e) {
                            console.error('Shadow DOM traversal error:', e);
                        }
                        return result;
                    }

                    const shadowHosts = [...document.querySelectorAll('*')].filter(el => el.shadowRoot);
                    const allShadowHTML = {};
                    shadowHosts.forEach((host, index) => {
                        const tagKey = host.tagName.toLowerCase() + '_' + index;
                        allShadowHTML[tagKey] = getShadowHTML(host);
                    });
                    return allShadowHTML;
                """
                shadow_html = driver.execute_script(shadow_js_script)
                
                shadow_log_path = os.path.join(DEBUG_DIR, f"{urlparse(url).path.replace('/', '_') or 'index'}_shadow_dom.json")
                with open(shadow_log_path, "w", encoding="utf-8") as f:
                    json.dump(shadow_html, f, indent=2, ensure_ascii=False)
                custom_print(f"📝 Shadow DOM content saved to {shadow_log_path}")
                
                try:
                    for host_tag, shadow_elements in shadow_html.items():
                        if not shadow_elements:
                            continue
                        for elem_html in shadow_elements:
                            shadow_soup = BeautifulSoup(elem_html, "lxml")
                            shadow_text = shadow_soup.get_text(" ", strip=True)
                            
                            debug_data["shadow_content"].append({
                                "host_tag": host_tag,
                                "html": elem_html,
                                "text": shadow_text[:1000]
                            })
                            
                            for match in EMAIL_REGEX.findall(shadow_text + " " + elem_html):
                                email = match.replace("(at)", "@").replace("(dot)", ".").replace("%40", "@").replace(" ", "")
                                email = email.replace('"email":"', "").replace('"', "")
                                if "@" in email and "." in email and not any(x in email for x in ["sentry", "wixpress"]):
                                    debug_data["raw_emails"].append(email)
                                    consolidated_links.add((email, "Shadow DOM Email"))
                            
                            for match in PHONE_REGEX.findall(shadow_text + " " + elem_html):
                                phone = re.sub(r'^(Tel|Phone|Call):\s*', '', match).strip()
                                phone = phone.replace("tel:", "").strip()
                                if len(phone) >= 8 and re.match(r'[\d\s\-\+\(\)]{8,}', phone):
                                    debug_data["raw_phones"].append(phone)
                            
                            for match in ADDRESS_REGEX.findall(shadow_text):
                                addr = match.strip()
                                if len(addr.split()) > 3:
                                    debug_data["raw_addresses"].append(addr)
                            
                            for platform, regex in SOCIAL_REGEXES.items():
                                matches = regex.findall(elem_html)
                                for m in matches:
                                    debug_data["raw_socials"].append((platform, m))
                                    formatted_match = m if m.startswith(('http://', 'https://')) else f"https://{m}"
                                    consolidated_links.add((formatted_match, "Shadow DOM Social"))
                            
                            for a in shadow_soup.find_all("a", href=True):
                                href = urljoin(url, a["href"])
                                consolidated_links.add((href, "Shadow DOM Link"))
                
                except Exception as e:
                    custom_print(f"❌ Error processing shadow DOM: {e}")
                    debug_data["alerts"].append(f"Error processing shadow DOM: {str(e)}")
                
                try:
                    iframes = driver.find_elements(By.TAG_NAME, "iframe")
                    for iframe in iframes:
                        try:
                            driver.switch_to.frame(iframe)
                            WebDriverWait(driver, 3).until(
                                EC.presence_of_element_located((By.TAG_NAME, "body"))
                            )
                            iframe_html = driver.page_source
                            debug_data["iframes"].append(iframe_html)
                            iframe_soup = BeautifulSoup(iframe_html, "lxml")
                            iframe_text = iframe_soup.get_text(" ", strip=True)

                            for match in EMAIL_REGEX.findall(iframe_text):
                                email = match.replace("(at)", "@").replace("(dot)", ".").replace("%40", "@").replace(" ", "")
                                email = email.replace('"email":"', "").replace('"', "")
                                if "@" in email and "." in email and not any(x in email for x in ["sentry", "wixpress"]):
                                    debug_data["raw_emails"].append(email)
                                    consolidated_links.add((email, "Iframe Email"))

                            for match in PHONE_REGEX.findall(iframe_text):
                                phone = re.sub(r'^(Tel|Phone|Call):\s*', '', match).strip()
                                if len(phone) >= 8 and re.match(r'[\d\s\-\+\(\)]{8,}', phone):
                                    debug_data["raw_phones"].append(phone)

                            for match in ADDRESS_REGEX.findall(iframe_text):
                                addr = match.strip()
                                if len(addr.split()) > 3:
                                    debug_data["raw_addresses"].append(addr)

                            for platform, regex in SOCIAL_REGEXES.items():
                                matches = regex.findall(iframe_html)
                                debug_data["raw_socials"].extend([(platform, m) for m in matches])
                                for m in matches:
                                    formatted_match = m if m.startswith(('http://', 'https://')) else f"https://{m}"
                                    consolidated_links.add((formatted_match, "Iframe Social"))

                            for a in iframe_soup.find_all("a", href=True):
                                href = urljoin(url, a["href"])
                                consolidated_links.add((href, "Iframe"))

                            driver.switch_to.default_content()
                        except Exception as e:
                            custom_print(f"Error processing iframe: {e}")
                            driver.switch_to.default_content()
                except:
                    pass
                
                # Get final page source
                html = driver.page_source
                debug_data["dynamic_html"] = html
                
                custom_print(f"📧 Shadow emails found: {list(set(debug_data['raw_emails']))}")
                custom_print(f"📞 Shadow phones found: {list(set(debug_data['raw_phones']))}")
                custom_print(f"🏠 Shadow addresses found: {list(set(debug_data['raw_addresses']))}")
                custom_print(f"🌐 Shadow socials found: {debug_data['raw_socials']}")
                
                return html, debug_data
                
            except Exception as e:
                error_detail = {
                    "method": method_name,
                    "attempt": attempt + 1,
                    "error_type": type(e).__name__,
                    "message": str(e),
                    "http_status": None,
                    "current_url": driver.current_url if hasattr(driver, 'current_url') else "Unknown"
                }
                try:
                    response = requests.head(url, timeout=10, allow_redirects=True)
                    error_detail["http_status"] = response.status_code
                except Exception as req_e:
                    custom_print(f"❌ Failed to get HTTP status for {url}: {req_e}")
                
                debug_data["error_details"].append(error_detail)
                debug_data["alerts"].append(f"{method_name}_attempt_{attempt + 1}_error: {str(e)}")
                custom_print(f"❌ {method_name} method failed for {url} (attempt {attempt + 1}): {error_detail['error_type']}: {error_detail['message']} (HTTP Status: {error_detail['http_status']})")
                continue
        
        time.sleep(random.uniform(3, 7))
    
    custom_print(f"❌ All access methods failed for {url}")
    debug_data["dynamic_html"] = "All access methods failed"
    debug_data["status"] = "failed"
    return "", debug_data

def extract_structured_address(text):
    """Extract structured address information from text"""
    # Common address patterns
    address_patterns = [
        r'(\d+\s+[\w\s]+(?:\s+(?:St|Street|Ave|Avenue|Rd|Road|Blvd|Boulevard|Ln|Lane))?[\s,]+[\w\s,]+(?:\s+\d{5}(?:-\d{4})?)?)',
        r'([\w\s]+(?:\s+(?:St|Street|Ave|Avenue|Rd|Road|Blvd|Boulevard))?[\s,]+(?:Suite|Ste\.?|Unit|Apt\.?|#)?\s*\d*[\s,]+[\w\s,]+(?:\s+\d{5}(?:-\d{4})?)?)',
        r'([A-Za-z\s]+,\s*[A-Za-z\s]+,\s*[A-Za-z\s]+(?:\s+\d{5})?)',
    ]
    
    addresses = []
    for pattern in address_patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            if isinstance(match, tuple):
                match = match[0]
            addr = match.strip()
            # Check if it looks like a real address (not just random text)
            if (len(addr.split()) >= 4 and 
                any(keyword in addr.lower() for keyword in ['st', 'street', 'ave', 'avenue', 'rd', 'road', 'blvd', 'boulevard']) or
                re.search(r'\d{5}(?:-\d{4})?', addr)):
                addresses.append(addr)
    
    return addresses

# Extract location info with Llama
def extract_location_info(text):
    if not pipeline:
        custom_print("❌ Llama model not loaded, returning N/A for location info")
        return {"country": "N/A"}

    prompt = f"""From this text, extract ONLY the country where the company is located. If no country is mentioned, return "N/A".
    
Text: {text[:2000]}  # Limit text length
    
Return a JSON object with a "country" key. Example: {{"country": "United States"}} or {{"country": "N/A"}}
Return ONLY the JSON object, no additional text."""

    messages = [
        {"role": "system", "content": "You are a helpful assistant that extracts country information from text. Return ONLY valid JSON."},
        {"role": "user", "content": prompt}
    ]

    try:
        outputs = pipeline(
            messages,
            max_new_tokens=100,
            temperature=0.1,
        )
        content = outputs[0]["generated_text"][-1]["content"]
        
        # Extract JSON
        json_match = re.search(r'\{.*\}', content, re.DOTALL)
        if json_match:
            result = json.loads(json_match.group(0))
            country = result.get("country", "N/A")
            custom_print(f"🌍 Extracted country: {country}")
            return {"country": country}
        
        return {"country": "N/A"}
        
    except Exception as e:
        custom_print(f"❌ Error extracting country info: {e}")
        return {"country": "N/A"}

# Extract information from HTML
def scrape_info(html, is_about_page=False, is_contact_page=False, url="unknown", debug_data=None):
    soup = BeautifulSoup(html, "lxml")
    text = soup.get_text(" ", strip=True)
    
    # Initialize sets and lists for collecting data
    emails = set()
    phones = set()
    addresses = set()
    socials = {platform: [] for platform in SOCIAL_REGEXES.keys()}
    all_hrefs = []  # List to store all <a href> tags with metadata
    footer_tracking = []
    
    # Helper function to verify and categorize hrefs
    def verify_href(href, source):
        href_clean = href.strip('"\' #').rstrip('/')  # Preserve query parameters, only strip fragment and quotes
        href_type = "Link"  # Default type
        
        # Check for email (mailto: links)
        if href.startswith("mailto:"):
            email = href.replace("mailto:", "").split("?")[0]
            if EMAIL_REGEX.match(email) and not any(x in email for x in ["sentry", "wixpress"]):
                emails.add(email)
                footer_tracking.append(f"Email found in {source} <a href>: {email}")
                custom_print(f"✅ Verified {source} email: {email}")
                return {"url": href, "source": source, "type": "Email", "page_url": url}
        
        # Check for social media links with validation
        for platform, regex in SOCIAL_REGEXES.items():
            if regex.search(href_clean):
                validated_link = validate_social_link(href_clean, platform)
                if validated_link:
                    if validated_link not in socials[platform]:
                        socials[platform].append(validated_link)
                        footer_tracking.append(f"Social {platform} found in {source} <a href>: {validated_link}")
                        custom_print(f"✅ Verified {source} {platform} link: {validated_link}")
                    return {"url": validated_link, "source": source, "type": f"Social ({platform})", "page_url": url}
                else:
                    custom_print(f"🚫 Rejected {platform} media/invalid URL: {href_clean}")
        
        # Check for phone (tel: links)
        if href.startswith("tel:"):
            phone = href.replace("tel:", "").strip()
            if re.match(r'^\+\d+', phone) and len(phone) >= 8:
                phones.add(phone)
                footer_tracking.append(f"Phone found in {source} <a href>: {phone}")
                custom_print(f"✅ Verified {source} phone: {phone}")
                return {"url": href, "source": source, "type": "Phone", "page_url": url}
        
        # Standard link (preserve query parameters)
        return {"url": urljoin(url, href).split("#")[0], "source": source, "type": "Link", "page_url": url}

    # Collect all <a href> tags from the entire page
    excluded_extensions = ['.pdf', '.jpg', '.png', '.jpeg', '.gif']
    page_links = [a["href"] for a in soup.select('a[href]') if a["href"] and not any(a["href"].lower().endswith(ext) for ext in excluded_extensions)]
    for href in page_links:
        href = urljoin(url, href).split("#")[0]  # Only remove fragment, preserve query parameters
        if href:  # Filter out empty hrefs
            all_hrefs.append(verify_href(href, "Page"))

    # Integrate shadow DOM data
    if debug_data:
        custom_print("🔄 Integrating shadow DOM data...")
        
        # Integrate shadow DOM emails
        if debug_data.get("raw_emails"):
            for email in debug_data["raw_emails"]:
                clean_email = email.replace("(at)", "@").replace("(dot)", ".").replace("%40", "@").replace(" ", "")
                clean_email = clean_email.replace('"email":"', "").replace('"', "")
                if "@" in clean_email and "." in clean_email and not any(x in clean_email for x in ["sentry", "wixpress"]):
                    emails.add(clean_email)
                    footer_tracking.append(f"Email found in shadow DOM: {clean_email}")
                    custom_print(f"✅ Added shadow DOM email: {clean_email}")
        
        # Integrate shadow DOM phones
        if debug_data.get("raw_phones"):
            for phone in debug_data["raw_phones"]:
                clean_phone = re.sub(r'^(Tel|Phone|Call):\s*', '', phone).strip()
                clean_phone = clean_phone.replace("tel:", "").strip()
                if len(clean_phone) >= 8 and re.match(r'[\d\s\-\+\(\)]{8,}', clean_phone):
                    phones.add(clean_phone)
                    footer_tracking.append(f"Phone found in shadow DOM: {clean_phone}")
                    custom_print(f"✅ Added shadow DOM phone: {clean_phone}")
        
        # Integrate shadow DOM addresses
        if debug_data.get("raw_addresses"):
            for addr in debug_data["raw_addresses"]:
                clean_addr = addr.strip()
                if len(clean_addr.split()) > 3:
                    addresses.add(clean_addr)
                    footer_tracking.append(f"Address found in shadow DOM: {clean_addr}")
                    custom_print(f"✅ Added shadow DOM address: {clean_addr}")
        
        # Integrate shadow DOM social media links with validation
        if debug_data.get("shadow_content"):
            for shadow_item in debug_data["shadow_content"]:
                shadow_soup = BeautifulSoup(shadow_item["html"], "lxml")
                
                # Process shadow DOM HTML for social links
                for platform, regex in SOCIAL_REGEXES.items():
                    matches = regex.findall(shadow_item["html"])
                    for match in matches:
                        cleaned_match = match.strip('"\' #').rstrip('/')
                        if not cleaned_match.startswith(('http://', 'https://')):
                            cleaned_match = f"https://{cleaned_match}"
                        
                        validated_link = validate_social_link(cleaned_match, platform)
                        if validated_link and validated_link not in socials[platform]:
                            socials[platform].append(validated_link)
                            footer_tracking.append(f"Social {platform} found in shadow DOM: {validated_link}")
                            custom_print(f"✅ Added shadow DOM {platform} link: {validated_link}")
                        elif not validated_link:
                            custom_print(f"🚫 Rejected shadow DOM {platform} media/invalid URL: {cleaned_match}")
                
                # Process shadow DOM <a> tags
                for a in shadow_soup.find_all("a", href=True):
                    href = urljoin(url, a["href"]).split("#")[0]  # Only remove fragment
                    if href:
                        all_hrefs.append(verify_href(href, "Shadow DOM"))

    # Enhanced footer detection
    footer_selector = (
        'footer, *[tagName*="footer" i], [class*="footer" i], [id*="footer" i], '
        '[role="contentinfo"], explorug-footer, explorug-contact, #root explorug-footer, '
        '#root explorug-contact, #root [class*="footer" i], #root [class*="Footer" i], '
        '[class*="bottom" i], [class*="foot" i], #root > *:last-child, '
        '[class*="social" i], [class*="list-social" i]'
    )
    footer = soup.select_one(footer_selector)
    
    footer_html = str(footer) if footer else "No footer found"
    footer_text = footer.get_text(" ", strip=True) if footer else ""
    
    # Extract footerURL attribute if present
    footer_urls = []
    if footer and footer.get("footerURL"):
        try:
            footer_url_data = json.loads(footer.get("footerURL").replace("'", "\""))
            footer_urls = [
                {"key": key, "url": url} for key, url in footer_url_data.items()
                if url.startswith(('http://', 'https://'))
            ]
            for item in footer_urls:
                all_hrefs.append(verify_href(item["url"], "Footer footerURL"))
        except json.JSONDecodeError:
            custom_print(f"❌ Failed to parse footerURL attribute: {footer.get('footerURL')}")
    
    # Collect <a href> tags from footer
    footer_nav_links = []  # Store footer navigation links for Llama processing
    if footer:
        for href_elem in footer.select('a[href]'):
            if href_elem["href"] and not any(href_elem["href"].lower().endswith(ext) for ext in excluded_extensions):
                href = urljoin(url, href_elem["href"]).split("#")[0]
                link_text = href_elem.get_text().strip()
                if href:
                    all_hrefs.append(verify_href(href_elem["href"], "Footer"))
                    # Store for Llama processing (only http/https links)
                    if href.startswith(('http://', 'https://')):
                        footer_nav_links.append({"url": href, "text": link_text})

    # Collect navbar/header links
    navbar_links = []
    navbar_selector = (
        'nav, header, [role="navigation"], [class*="nav" i], [class*="menu" i], '
        '[class*="header" i], [id*="nav" i], [id*="menu" i], [id*="header" i]'
    )
    navbar = soup.select(navbar_selector)
    for nav_elem in navbar:
        for href_elem in nav_elem.select('a[href]'):
            if href_elem["href"] and not any(href_elem["href"].lower().endswith(ext) for ext in excluded_extensions):
                href = urljoin(url, href_elem["href"]).split("#")[0]
                link_text = href_elem.get_text().strip()
                if href.startswith(('http://', 'https://')):
                    navbar_links.append({"url": href, "text": link_text})

    custom_print(f"📋 Collected {len(navbar_links)} navbar links and {len(footer_nav_links)} footer links")

    # Process page sections
    about_sections = []
    contact_sections = []
    contact_text = ""
    
    if is_about_page:
        about_sections = soup.select('div[class*="about" i], section[class*="about" i], article[class*="about" i], '
                                   'div[id*="about" i], section[id*="about" i], article[id*="about" i], '
                                   'div[class*="company" i], div[class*="mission" i], div[class*="vision" i], '
                                   'div[id*="company" i], div[id*="mission" i], div[id*="vision" i]')
        if not about_sections:
            about_sections = [soup.select_one('main, [role="main"], body')]

    if is_contact_page:
        contact_sections = soup.select('div[class*="contact" i], section[class*="contact" i], article[class*="contact" i], '
                                     'div[id*="contact" i], section[id*="contact" i], article[id*="contact" i], '
                                     'div[class*="reach-us" i], div[class*="get-in-touch" i], '
                                     'div[id*="reach-us" i], div[id*="get-in-touch" i], '
                                     'explorug-contact')
        if not contact_sections:
            contact_sections = [soup.select_one('main, [role="main"], body')]
        
        for section in contact_sections:
            if section:
                section_text = section.get_text(" ", strip=True)
                if section_text:
                    contact_text += section_text + " "

    # Regular DOM email extraction from text (non-href sources)
    for match in EMAIL_REGEX.findall(text):
        email = match.replace("(at)", "@").replace("(dot)", ".").replace("%40", "@").replace(" ", "")
        email = email.replace('"email":"', "").replace('"', "")
        if "@" in email and "." in email and not any(x in email for x in ["sentry", "wixpress"]):
            if email not in emails:
                emails.add(email)
                footer_tracking.append(f"Email found in page text: {email}")

    if footer_text:
        for match in EMAIL_REGEX.findall(footer_text):
            email = match.replace("(at)", "@").replace("(dot)", ".").replace("%40", "@").replace(" ", "")
            email = email.replace('"email":"', "").replace('"', "")
            if "@" in email and "." in email and not any(x in email for x in ["sentry", "wixpress"]):
                if email not in emails:
                    emails.add(email)
                    footer_tracking.append(f"Email found in footer text: {email}")

    # Regular DOM social media extraction from HTML (non-href sources) with validation
    for platform, regex in SOCIAL_REGEXES.items():
        matches = regex.findall(html)
        for match in matches:
            cleaned_match = match.strip('"\' ?#').rstrip('/')
            if not cleaned_match.startswith(('http://', 'https://')):
                cleaned_match = f"https://{cleaned_match}"
            
            validated_link = validate_social_link(cleaned_match, platform)
            if validated_link and validated_link not in socials[platform]:
                socials[platform].append(validated_link)
                footer_tracking.append(f"Social {platform} found in page HTML (non-href): {validated_link}")
                custom_print(f"✅ Added page {platform} link (non-href): {validated_link}")
            elif not validated_link:
                custom_print(f"🚫 Rejected page {platform} media/invalid URL: {cleaned_match}")

    # Regular DOM phone extraction from text
    if footer_text:
        for match in PHONE_REGEX.findall(footer_text):
            phone = match[0] if match[0] else f"{match[1]}{match[2]}{match[3]}"
            phone = phone.replace("tel:", "").strip()
            if len(phone) >= 8 and phone not in phones:
                phones.add(phone)

        for match in PHONE_REGEX.findall(footer_text):
            phone = match[0] if match[0] else f"{match[1]}{match[2]}{match[3]}"
            if re.match(r'^\+\d+', phone) and len(phone) >= 8 and phone not in phones:
                phones.add(phone)

    # Regular DOM address extraction
    if footer_text:
        # Use both regex and structured extraction
        for match in ADDRESS_REGEX.findall(footer_text):
            addr = match.strip()
            if len(addr.split()) > 3 and addr not in addresses:
                addresses.add(addr)
        
        # Use structured extraction
        structured_addresses = extract_structured_address(footer_text)
        for addr in structured_addresses:
            addresses.add(addr)

    # Extract from main text too
    for match in ADDRESS_REGEX.findall(text):
        addr = match.strip()
        if len(addr.split()) > 3 and addr not in addresses:
            addresses.add(addr)
    
    structured_addresses = extract_structured_address(text)
    for addr in structured_addresses:
        addresses.add(addr)

    # Clean up social media lists and remove duplicates
    for platform in socials:
        socials[platform] = sorted(list(set(socials[platform])))
        custom_print(f"Final {platform} links: {socials[platform]}")

    # Extract meta description
    desc = ""
    desc_tags = soup.select('meta[name*="description" i], meta[property*="og:description" i], meta[name*="twitter:description" i]')
    for tag in desc_tags:
        if tag and "content" in tag.attrs:
            tag_desc = tag["content"].strip()
            if tag.attrs.get("name", "").lower() == "description" and tag_desc:
                desc = tag_desc
                break
            elif len(tag_desc) > len(desc):
                desc = tag_desc

    # Extract about content
    about_content = ""
    if is_about_page:
        for section in about_sections:
            if section:
                section_text = section.get_text(" ", strip=True)
                if section_text and len(section_text) > len(about_content):
                    about_content = section_text[:2000]

    combined_text = (footer_text + " " + contact_text).strip()

    # Save tracking information
    footer_tracking_path = os.path.join(DEBUG_DIR, f"{urlparse(url).path.replace('/', '_') or 'index'}_footer_tracking.txt")
    with open(footer_tracking_path, "w", encoding="utf-8") as f:
        f.write(f"URL: {url}\n")
        if footer_tracking:
            f.write("\n".join(footer_tracking) + "\n")
        else:
            f.write("No data extraction tracking available.\n")

    # Save all hrefs to debug file
    hrefs_log_path = os.path.join(DEBUG_DIR, f"{urlparse(url).path.replace('/', '_') or 'index'}_hrefs.json")
    with open(hrefs_log_path, "w", encoding="utf-8") as f:
        json.dump(all_hrefs, f, indent=2, ensure_ascii=False)
    custom_print(f"📝 All hrefs saved to {hrefs_log_path}")

    # Validate phone numbers with country codes
    validated_phones = []
    for phone in phones:
        validation_result = validate_phone_with_country_code(phone)
        if validation_result["is_valid"]:
            validated_phones.append(validation_result)
            custom_print(f"✅ Validated phone: {phone} -> {validation_result['country']} ({validation_result['country_code']})")
        else:
            # Keep unvalidated phones too, but mark them
            validated_phones.append(validation_result)
            custom_print(f"⚠️ Phone without valid country code: {phone}")

    custom_print(f"📧 Final emails: {list(emails)}")
    custom_print(f"📞 Final phones: {len(validated_phones)} ({len([p for p in validated_phones if p['is_valid']])} validated)")
    custom_print(f"🏠 Final addresses: {list(addresses)}")
    custom_print(f"🔗 Total hrefs collected: {len(all_hrefs)}")

    return (
        emails,
        socials,
        desc,
        about_content,
        validated_phones,  # Return validated phones instead of raw phones
        combined_text,
        footer_html,
        footer_text,
        addresses,
        all_hrefs,
        navbar_links,
        footer_nav_links
    )

# Generate business nature with Llama
def generate_business_nature(meta_desc, about_content):
    if not pipeline:
        custom_print("❌ Llama model not loaded, returning N/A for business nature")
        return "N/A (Llama model not loaded)"

    # Use footer text or combined text if about_content is empty
    input_text = ""
    if meta_desc:
        input_text += f"Meta Description: {meta_desc}\n"
    if about_content:
        input_text += f"About Us: {about_content[:1500]}"
    
    # If no proper inputs, return a generic message
    if len(input_text.strip()) < 50:
        custom_print("⚠️ Insufficient content for business nature generation")
        return "N/A (Insufficient company information)"
    
    prompt = f"""Based on the following company information, describe the business nature in one concise sentence. Focus on what the company does, its industry, and its main products/services.

{input_text}

Describe the business nature concisely. Example: "A technology company specializing in software development for e-commerce platforms."
Return only the description, no additional text or explanations."""

    messages = [
        {"role": "system", "content": "You are a business analyst that describes company business nature based on available information."},
        {"role": "user", "content": prompt}
    ]

    try:
        outputs = pipeline(
            messages,
            max_new_tokens=100,
            temperature=0.1,
        )
        content = outputs[0]["generated_text"][-1]["content"].strip()
        
        # Clean up the response
        if content.startswith('"') and content.endswith('"'):
            content = content[1:-1]
        
        # Remove any prefix like "Business nature:" or similar
        content = re.sub(r'^(?:Business\s*Nature|Description|Company\s*Description):\s*', '', content, flags=re.IGNORECASE)

        with open(os.path.join(DEBUG_DIR, "llama_business_nature_output.txt"), "a", encoding="utf-8") as f:
            f.write(f"Input: {input_text}\nLlama output: {content}\n\n")

        custom_print(f"🏢 Generated business nature: {content}")
        return content
    except Exception as e:
        custom_print(f"❌ Error generating business nature with Llama: {e}")
        return "N/A (Error in Llama generation)"
    
def check_with_requests_session(url):
    """Check accessibility using requests with session and headers"""
    session = requests.Session()
    
    # Set up retry strategy
    retry_strategy = Retry(
        total=3,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504],
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    
    headers = {
        'User-Agent': random.choice(USER_AGENTS),
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br',
        'DNT': '1',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Cache-Control': 'max-age=0',
    }
    
    try:
        response = session.get(url, headers=headers, timeout=30, allow_redirects=True)
        return response.status_code in [200, 301, 302]
    except:
        return False

def check_with_selenium_quick(url):
    """Quick check using Selenium with enhanced options"""
    driver = None
    try:
        driver = init_selenium_driver()
        driver.get(url)
        # Wait a bit for potential redirects or dynamic loading
        time.sleep(3)
        
        # Check if page loaded successfully
        page_source = driver.page_source
        return len(page_source) > 1000 and "error" not in page_source.lower()[:500]
    
    except Exception:
        return False
    finally:
        if driver:
            driver.quit()

def check_with_curl(url):
    """Fallback using curl command"""
    try:
        cmd = [
            'curl', '-s', '-I', '--max-time', '30',
            '-H', f'User-Agent: {random.choice(USER_AGENTS)}',
            '-H', 'Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            url
        ]
        
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=35)
        return result.returncode == 0 and ('200 OK' in result.stdout or '301' in result.stdout or '302' in result.stdout)
    except:
        return False

# Check if a link is accessible
def is_link_accessible(url, max_retries=3):
    """Enhanced accessibility check with multiple methods and retry logic"""
    
    methods = [
        ('requests_session', check_with_requests_session),
        ('selenium_quick', check_with_selenium_quick),
        ('curl_fallback', check_with_curl)
    ]
    
    for method_name, method_func in methods:
        for attempt in range(max_retries):
            try:
                custom_print(f"🔄 Checking {url} with {method_name} (attempt {attempt + 1})")
                result = method_func(url)
                if result:
                    custom_print(f"✅ Link accessible via {method_name}: {url}")
                    return True
                time.sleep(random.uniform(1, 3))  # Random delay between attempts
            except Exception as e:
                custom_print(f"❌ {method_name} failed for {url}: {str(e)}")
                continue
    
    custom_print(f"⏭️ All methods failed for: {url}")
    return False

# Process page
def process_page(url, depth, parsed_domain, visited):
    if depth > MAX_DEPTH or url in visited or not urlparse(url).netloc.endswith(parsed_domain):
        custom_print(f"⏭️ Skipping {url} (depth: {depth}, visited: {url in visited}, domain: {urlparse(url).netloc})")
        return None, {
            "url": url,
            "status": "skipped",
            "dynamic_html": "",
            "links": [],
            "scripts": [],
            "raw_socials": [],
            "raw_emails": [],
            "raw_phones": [],
            "raw_addresses": [],
            "shadow_links": [],
            "shadow_content": [],
            "alerts": [],
            "consolidated_links": [],
            "all_links": [],
            "iframes": [],
            "error_details": []
        }, []

    if not is_link_accessible(url):
        custom_print(f"⏭️ Skipping inaccessible link: {url}")
        return None, {
            "url": url,
            "status": "skipped_inaccessible",
            "dynamic_html": "",
            "links": [],
            "scripts": [],
            "raw_socials": [],
            "raw_emails": [],
            "raw_phones": [],
            "raw_addresses": [],
            "shadow_links": [],
            "shadow_content": [],
            "alerts": [],
            "consolidated_links": [],
            "all_links": [],
            "iframes": [],
            "error_details": [{"error_type": "Inaccessible", "message": "Link failed accessibility checks"}]
        }, []

    visited.add(url)
    custom_print(f"🔎 Crawling: {url}")
    driver = init_selenium_driver()
    try:
        html, page_debug_data = get_rendered_html(url, driver)
        
        if not html or "blocked" in html.lower() or len(html) < 500:
            custom_print(f"❌ Selenium failed or returned blocked/minimal content for {url}. Falling back to primary requests.")
            page_debug_data["status"] = "fallback_requests"
            page_debug_data["alerts"].append("Selenium failed, using primary requests fallback")
            
            # Primary requests fallback
            session = requests.Session()
            headers = {
                'User-Agent': random.choice(USER_AGENTS),
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
                'Accept-Encoding': 'gzip, deflate, br',
                'DNT': '1',
                'Connection': 'keep-alive',
            }
            try:
                response = session.get(url, headers=headers, timeout=30, allow_redirects=True)
                if response.status_code == 200 and len(response.text) > 500 and "blocked" not in response.text.lower():
                    html = response.text
                    page_debug_data["dynamic_html"] = html
                    page_debug_data["alerts"].append("Successfully fetched HTML with primary requests fallback")
                    custom_print(f"✅ Primary requests fallback successful for {url}")
                else:
                    custom_print(f"❌ Primary requests fallback failed for {url}: Status {response.status_code} or blocked/minimal content")
                    page_debug_data["error_details"].append({
                        "method": "requests_fallback",
                        "attempt": 1,
                        "error_type": "FailedFallback",
                        "message": f"Primary requests returned status {response.status_code} or blocked/minimal content",
                        "http_status": response.status_code,
                        "current_url": url
                    })
                    # Second fallback: lightweight requests-based link collection
                    custom_print(f"❌ Primary requests fallback failed. Attempting lightweight requests fallback for link collection.")
                    page_debug_data["status"] = "fallback_lightweight_requests"
                    page_debug_data["alerts"].append("Primary requests failed, using lightweight requests fallback")
                    try:
                        headers = {
                            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36"
                        }
                        resp = requests.get(url, headers=headers, timeout=10)
                        resp.raise_for_status()
                        html = resp.text
                        page_debug_data["dynamic_html"] = html
                        page_debug_data["alerts"].append("Successfully fetched HTML with lightweight requests fallback")
                        custom_print(f"✅ Lightweight requests fallback successful for {url}")
                    except Exception as e:
                        custom_print(f"❌ Lightweight requests fallback failed for {url}: {str(e)}")
                        page_debug_data["error_details"].append({
                            "method": "lightweight_requests_fallback",
                            "attempt": 1,
                            "error_type": type(e).__name__,
                            "message": str(e),
                            "http_status": getattr(e.response, 'status_code', None),
                            "current_url": url
                        })
                        return None, page_debug_data, []
            except Exception as e:
                custom_print(f"❌ Primary requests fallback failed for {url}: {str(e)}")
                page_debug_data["error_details"].append({
                    "method": "requests_fallback",
                    "attempt": 1,
                    "error_type": type(e).__name__,
                    "message": str(e),
                    "http_status": getattr(e.response, 'status_code', None),
                    "current_url": url
                })
                # Second fallback: lightweight requests-based link collection
                custom_print(f"❌ Primary requests fallback failed. Attempting lightweight requests fallback for link collection.")
                page_debug_data["status"] = "fallback_lightweight_requests"
                page_debug_data["alerts"].append("Primary requests failed, using lightweight requests fallback")
                try:
                    headers = {
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/140.0.0.0 Safari/537.36"
                    }
                    resp = requests.get(url, headers=headers, timeout=10)
                    resp.raise_for_status()
                    html = resp.text
                    page_debug_data["dynamic_html"] = html
                    page_debug_data["alerts"].append("Successfully fetched HTML with lightweight requests fallback")
                    custom_print(f"✅ Lightweight requests fallback successful for {url}")
                except Exception as e:
                    custom_print(f"❌ Lightweight requests fallback failed for {url}: {str(e)}")
                    page_debug_data["error_details"].append({
                        "method": "lightweight_requests_fallback",
                        "attempt": 1,
                        "error_type": type(e).__name__,
                        "message": str(e),
                        "http_status": getattr(e.response, 'status_code', None),
                        "current_url": url
                    })
                    return None, page_debug_data, []

        is_about_page = "about" in url.lower()
        is_contact_page = "contact" in url.lower()

        emails, socials, desc, page_about_content, phones, combined_text, footer_html, footer_text, addresses, page_hrefs, navbar_links, footer_nav_links = scrape_info(
            html, is_about_page, is_contact_page, url, page_debug_data
        )

        custom_print(f"page_debug_data keys before update: {list(page_debug_data.keys())}")

        page_debug_data.update({
            "status": "crawled" if html == page_debug_data["dynamic_html"] and page_debug_data.get("status", "initial") not in ["fallback_requests", "fallback_lightweight_requests"] else page_debug_data.get("status", "initial"),
            "all_links": []
        })

        soup = BeautifulSoup(html, "lxml")
        discovered_links = []
        excluded_extensions = ['.pdf', '.jpg', '.png', '.jpeg', '.gif']
        for a in soup.select('a[href]:not([href$=".pdf"], [href$=".jpg"], [href$=".png"], [href$=".jpeg"], [href$=".gif"])'):
            next_url = urljoin(url, a["href"]).split("#")[0]
            parsed_next = urlparse(next_url)
            link_text = a.get_text().strip().lower()
            page_debug_data["all_links"].append({"href": next_url, "text": link_text})

        # Use Llama to intelligently identify About/Contact pages from navbar and footer links
        all_nav_links = navbar_links + footer_nav_links
        if all_nav_links and depth == 0:  # Only use Llama on the homepage to discover pages
            custom_print(f"🤖 Using Llama to identify About/Contact pages from {len(all_nav_links)} navigation links...")
            identified_links = identify_about_contact_links(all_nav_links)

            # Add identified About pages to crawl queue
            for link_info in identified_links.get("about", []):
                link_url = link_info["url"]
                parsed_link = urlparse(link_url)
                if (parsed_link.netloc.endswith(parsed_domain) and
                    link_url not in visited):
                    custom_print(f"📖 Llama identified About page: {link_url}")
                    discovered_links.append((link_url, depth + 1))

            # Add identified Contact pages to crawl queue
            for link_info in identified_links.get("contact", []):
                link_url = link_info["url"]
                parsed_link = urlparse(link_url)
                if (parsed_link.netloc.endswith(parsed_domain) and
                    link_url not in visited):
                    custom_print(f"📞 Llama identified Contact page: {link_url}")
                    discovered_links.append((link_url, depth + 1))
        else:
            # Fallback to keyword matching for non-homepage pages or when no nav links
            for a in soup.select('a[href]:not([href$=".pdf"], [href$=".jpg"], [href$=".png"], [href$=".jpeg"], [href$=".gif"])'):
                next_url = urljoin(url, a["href"]).split("#")[0]
                parsed_next = urlparse(next_url)
                link_text = a.get_text().strip().lower()

                if (parsed_next.netloc.endswith(parsed_domain) and
                    next_url not in visited and
                    ("about" in next_url.lower() or "contact" in next_url.lower() or "about" in link_text or "contact" in link_text)):
                    discovered_links.append((next_url, depth + 1))

        # Save extracted links to all_links.txt (similar to provided code)
        all_links_file = os.path.join(DEBUG_DIR, "all_links.txt")
        try:
            with open(all_links_file, "a", encoding="utf-8") as f:
                for link in page_hrefs:
                    if link["url"].startswith(('http://', 'https://', 'mailto:', 'tel:')):
                        f.write(f"{link['url']} (from {url})\n")
            custom_print(f"📝 Appended links to {all_links_file}")
        except Exception as e:
            custom_print(f"❌ Error saving links to {all_links_file}: {str(e)}")
            page_debug_data["error_details"].append({
                "method": "save_links",
                "attempt": 1,
                "error_type": type(e).__name__,
                "message": str(e),
                "http_status": None,
                "current_url": url
            })

        return (emails, socials, desc, page_about_content, phones, combined_text, footer_html, footer_text, addresses, page_hrefs), page_debug_data, discovered_links
    finally:
        driver.quit()

def crawl_website(domain, max_depth=MAX_DEPTH):
    visited = set()
    to_visit = [(domain, 0)]
    all_emails = set()
    all_socials = {}
    all_phones = set()
    all_addresses = set()
    all_links = set()
    all_footer_contact_texts = []
    meta_description = ""
    about_content = ""
    footer_contents = []
    debug_info = []
    all_hrefs_collection = []

    parsed_domain = urlparse(domain).netloc

    while to_visit:
        url, depth = to_visit.pop(0)
        
        result, page_debug_data, discovered_links = process_page(url, depth, parsed_domain, visited)
        
        if result:
            emails, socials, desc, page_about_content, phones, combined_text, footer_html, footer_text, addresses, page_hrefs = result
            all_emails.update(emails)
            
            for platform, links in socials.items():
                if platform not in all_socials:
                    all_socials[platform] = []
                all_socials[platform].extend([link for link in links if link not in all_socials.get(platform, [])])
            
            # Convert phone dictionaries to strings before adding to set
            for phone_data in phones:
                if isinstance(phone_data, dict):
                    all_phones.add(phone_data.get("phone", ""))
                else:
                    all_phones.add(str(phone_data))
            
            all_addresses.update(addresses)
            
            # FIX: Extract just the href string from the dictionary
            for link_dict in page_debug_data["all_links"]:
                if link_dict.get("href", "").startswith(('http://', 'https://')):
                    all_links.add(link_dict["href"])
            
            if combined_text:
                all_footer_contact_texts.append(combined_text)
            
            # FIX: Collect meta description and about content properly
            if not meta_description and desc:
                meta_description = desc
                custom_print(f"📝 Collected meta description: {desc[:100]}...")
            
            if page_about_content and len(page_about_content) > len(about_content):
                about_content = page_about_content
                custom_print(f"📝 Collected about content: {page_about_content[:100]}...")
            
            if footer_html != "No footer found":
                footer_contents.append({
                    "url": page_debug_data["url"],
                    "html": footer_html,
                    "text": footer_text
                })
            
            all_hrefs_collection.extend(page_hrefs)
            debug_info.append(page_debug_data)
        
        else:
            debug_info.append(page_debug_data)  # Include debug_data even for failed pages
        
        time.sleep(CRAWL_DELAY)

    all_hrefs_log_path = os.path.join(DEBUG_DIR, f"all_hrefs_{parsed_domain.replace('.', '_')}.json")
    try:
        with open(all_hrefs_log_path, "w", encoding="utf-8") as f:
            json.dump(all_hrefs_collection, f, indent=2, ensure_ascii=False)
        custom_print(f"📝 All collected <a href> links saved to {all_hrefs_log_path}")
    except Exception as e:
        custom_print(f"❌ Error saving all hrefs log file: {str(e)}")

    combined_text = " ".join(all_footer_contact_texts)
    
    # FIX: Extract country from addresses if available
    location_info = {"country": "N/A"}
    if addresses:
        # Try to extract country from addresses
        for address in addresses:
            if "Nepal" in address or "Kathmandu" in address:
                location_info["country"] = "Nepal"
                custom_print(f"🌍 Extracted country from address: Nepal")
                break
    
    # FIX: If no country found in addresses, use Llama
    if location_info["country"] == "N/A" and combined_text:
        location_info = extract_location_info(combined_text)
    
    # FIX: Debug output for business nature generation
    custom_print(f"📝 Business nature generation inputs:")
    custom_print(f"  Meta description: {'Yes' if meta_description else 'No'} ({len(meta_description)} chars)")
    custom_print(f"  About content: {'Yes' if about_content else 'No'} ({len(about_content)} chars)")
    
    # Try to extract from footer if no about content
    if not about_content and footer_contents:
        for fc in footer_contents:
            if "About" in fc["text"] or "about" in fc["text"]:
                about_content = fc["text"][:2000]
                custom_print(f"📝 Using footer text for about content")
                break
    
    business_nature = generate_business_nature(meta_description, about_content)
    
    return all_emails, all_socials, meta_description, about_content, all_phones, all_addresses, all_links, location_info, business_nature, footer_contents, debug_info

def save_results(domain, emails, socials, meta_description, about_content, phones, addresses, links, location_info, business_nature, footer_contents, debug_info, process_log, company_links=None, company_name="Unknown"):
    results = {
        "company_name": company_name,
        "domain": domain,
        "emails": list(emails),
        "social_media": socials,
        "meta_description": meta_description or "N/A",
        "about_content": about_content or "N/A",
        "phone_numbers": list(phones),
        "addresses": list(addresses),
        "links": list(links),
        "country": location_info.get("country", "N/A"),  # Only country now
        "business_nature": business_nature or "N/A",
        "company_links": company_links or [],
        "process_log": process_log
    }

    output_file = os.path.join(DEBUG_DIR, f"scrape_results_{company_name.replace(' ', '_')}.json")
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2)
    custom_print(f"📝 Core scraped information and process logs saved to {output_file}")

    with open(os.path.join(DEBUG_DIR, f"footer_content_{company_name.replace(' ', '_')}.json"), "w", encoding="utf-8") as f:
        json.dump([
            {
                "url": fc["url"],
                "html": fc["html"],
                "text": fc["text"]
            } for fc in footer_contents
        ], f, indent=2)
    footer_filename = f"footer_content_{company_name.replace(' ', '_')}.json"
    custom_print(f"📝 Footer content saved to {os.path.join(DEBUG_DIR, footer_filename)}")

    with open(os.path.join(DEBUG_DIR, f"debug_info_{company_name.replace(' ', '_')}.json"), "w", encoding="utf-8") as f:
        json.dump([
            {
                "url": di["url"],
                "status": di.get("status", "unknown"),
                "dynamic_html": di.get("dynamic_html", ""),
                "links": di.get("links", []),
                "scripts": di.get("scripts", []),
                "raw_socials": di.get("raw_socials", []),
                "raw_emails": di.get("raw_emails", []),
                "raw_phones": di.get("raw_phones", []),
                "raw_addresses": di.get("raw_addresses", []),
                "shadow_links": di.get("shadow_links", []),
                "shadow_content": di.get("shadow_content", []),
                "footer_tracking": di.get("footer_tracking", []),
                "meta_tags": di.get("meta_tags", []),
                "page_content": di.get("page_content", {}),
                "all_links": di.get("all_links", []),
                "consolidated_links": di.get("consolidated_links", []),
                "fallback_logs": di.get("fallback_logs", []),
                "alerts": di.get("alerts", []),
                "iframes": di.get("iframes", [])
            } for di in debug_info
        ], f, indent=2)
    debug_filename = f"debug_info_{company_name.replace(' ', '_')}.json"
    custom_print(f"📝 Debug information saved to {os.path.join(DEBUG_DIR, debug_filename)}")

    return results

# Fallback Google search
def fallback_google_search(query, company, location, social_platforms=None):
    debug_log = f"Query: {query}\nResults:\n"
    socials = {platform: [] for platform in social_platforms} if social_platforms else set()
    company_links = []
    company_clean = re.sub(r'\s*(pvt|ltd|limited|private|inc|incorporated|co|company)\s*$', '', company, flags=re.IGNORECASE).lower()
    location_lower = location.lower() if location else ""
    company_variations = [company_clean, company_clean.replace(" ", ""), company_clean.replace(" ", "-")]
    location_variations = [location_lower, location_lower.replace(" ", ""), location_lower.replace(" ", "-")] if location else [""]
    excluded_terms = ["login", "security", "plugins", "help", "signup", "terms", "privacy", "ads", "careers", "policies", "support"]

    for api_key in GOOGLE_API_KEYS:
        try:
            service = build("customsearch", "v1", developerKey=api_key)
            res = service.cse().list(q=query, cx=SEARCH_ENGINE_ID, num=10).execute()
            results = res.get("items", [])
            
            for item in results:
                title = item.get("title", "").lower()
                snippet = item.get("snippet", "").lower()
                link = item.get("link", "")
                link_lower = link.lower()
                path = urlparse(link).path.lower().strip('/')
                debug_log += f"Title: {title}\nLink: {link}\nSnippet: {snippet}\n"
                
                is_company_relevant = False
                for comp_var in company_variations:
                    if (comp_var in title or
                         comp_var in snippet or
                         comp_var in link_lower or
                        (location and f"{comp_var}{location_variations[0]}" in title) or
                        (location and f"{comp_var}{location_variations[0]}" in snippet) or
                        (location and f"{comp_var}{location_variations[0]}" in link_lower)):
                        is_company_relevant = True
                        break
                
                is_location_relevant = any(loc_var in title or loc_var in snippet or loc_var in link_lower for loc_var in location_variations) if location else True
                context_indicators = ["official", "profile", "page", "account", company_clean] + ([location_lower] if location else [])
                is_context_relevant = any(indicator in title or indicator in snippet for indicator in context_indicators)
                
                if social_platforms and is_company_relevant and is_context_relevant:
                    for platform in social_platforms:
                        regex = SOCIAL_REGEXES.get(platform)
                        if regex and regex.search(link):
                            # Apply validation to fallback results too
                            validated_link = validate_social_link(link, platform)
                            if validated_link and not any(x in validated_link.lower() for x in excluded_terms):
                                if validated_link not in socials[platform]:
                                    socials[platform].append(validated_link)
                                    debug_log += (
                                        f"Accepted {platform} link: {validated_link}\n"
                                        f"Path: {path}\n"
                                        f"Company relevance: {any(comp_var in path for comp_var in company_variations)}\n"
                                        f"Location relevance: {any(loc_var in path for loc_var in location_variations) if location else 'N/A'}\n"
                                        f"Context relevance: {is_context_relevant}\n---\n"
                                    )
                            else:
                                debug_log += f"Rejected {platform} link: {link} (failed validation or contains excluded terms)\n---\n"
                        else:
                            debug_log += f"Rejected link for {platform}: {link} (no regex match)\n---\n"
                
                if is_company_relevant:
                    formatted_link = link if link.startswith(('http://', 'https://')) else f"https://{link}"
                    if (not any(x in formatted_link.lower() for x in excluded_terms) and
                        not any(formatted_link in socials[platform] for platform in socials)):
                        parsed_link = urlparse(formatted_link)
                        if parsed_link.netloc:
                            company_links.append(formatted_link)
                            debug_log += (
                                f"Accepted company link: {formatted_link}\n"
                                f"Path: {path}\n"
                                f"Company relevance: {is_company_relevant}\n"
                                f"Location relevance: {is_location_relevant}\n"
                                f"Context relevance: {is_context_relevant}\n---\n"
                            )
                        else:
                            debug_log += f"Rejected company link: {formatted_link} (invalid domain)\n---\n"
                    else:
                        debug_log += f"Rejected company link: {formatted_link} (contains excluded terms or already in socials)\n---\n"
                else:
                    debug_log += f"Rejected company link: {link} (not company relevant)\n---\n"
            
            for platform in socials:
                socials[platform] = sorted(list(set(socials[platform])))
            company_links = sorted(list(set(company_links)))
            
            with open(os.path.join(DEBUG_DIR, f"fallback_search_{query.replace(' ', '_')}.txt"), "a", encoding="utf-8") as f:
                f.write(debug_log + "\n")
            custom_print(f"🔍 Fallback search for {query}: Found {sum(len(v) for v in socials.values()) if social_platforms else len(socials)} {'links' if social_platforms else 'emails'}, {len(company_links)} company links")
            return socials, company_links, debug_log
            
        except HttpError as e:
            if e.resp.status == 429:
                custom_print(f"❌ Google API error with key {api_key}: Quota exceeded. Trying next API key...")
                debug_log += f"Error with key {api_key}: Quota exceeded\n"
                continue
            else:
                custom_print(f"❌ Fallback Google API error with key {api_key} for query '{query}': {e}")
                debug_log += f"Error with key {api_key}: {e}\n"
                with open(os.path.join(DEBUG_DIR, f"fallback_search_{query.replace(' ', '_')}.txt"), "a", encoding="utf-8") as f:
                    f.write(debug_log + "\n")
                return ({} if social_platforms else set()), [], debug_log
        except Exception as e:
            custom_print(f"❌ Fallback Google API error with key {api_key} for query '{query}': {e}")
            debug_log += f"Error with key {api_key}: {e}\n"
            with open(os.path.join(DEBUG_DIR, f"fallback_search_{query.replace(' ', '_')}.txt"), "a", encoding="utf-8") as f:
                f.write(debug_log + "\n")
            return ({} if social_platforms else set()), [], debug_log
    
    custom_print(f"❌ All API keys failed or quota exceeded for query '{query}'.")
    debug_log += "All API keys failed or quota exceeded\n"
    with open(os.path.join(DEBUG_DIR, f"fallback_search_{query.replace(' ', '_')}.txt"), "a", encoding="utf-8") as f:
        f.write(debug_log + "\n")
    return ({} if social_platforms else set()), [], debug_log

def validate_company_match(domain, company_name):
    """Validate if a domain belongs to the given company with strict matching"""
    if not domain or not company_name:
        return False
    
    parsed = urlparse(domain)
    domain_name = parsed.netloc.lower()
    
    # Clean company name
    company_clean = re.sub(r'\s*(pvt|ltd|limited|private|inc|incorporated|co|company|corp|corporation|llc|gmbh|plc)\s*\.?$', 
                          '', company_name, flags=re.IGNORECASE)
    company_clean = company_clean.lower().strip()
    
    # Get company words (ignore common short words)
    company_words = [word for word in re.findall(r'\b\w+\b', company_clean) if len(word) > 3]
    
    if not company_words:
        company_words = re.findall(r'\b\w+\b', company_clean)
    
    # Extract domain without TLD and www
    domain_base = re.sub(r'^www\.', '', domain_name)
    domain_base = re.sub(r'\.(com|org|net|io|co|in|us|uk|au|ca|de|fr|jp|biz|info|mobi|name|tv|cc|ws)$', '', domain_base)
    
    # Check if domain contains company words
    matches = []
    for word in company_words:
        if word in domain_base:
            matches.append(word)
    
    # Calculate match percentage
    if company_words:
        match_percentage = (len(matches) / len(company_words)) * 100
        
        # Strict criteria: At least 50% match OR first word must match
        if match_percentage >= 50 or (company_words and company_words[0] in domain_base):
            custom_print(f"✅ Company match: {match_percentage:.0f}% - Words matched: {matches}")
            return True
        else:
            custom_print(f"❌ Weak company match: {match_percentage:.0f}% - Only matched: {matches}")
            return False
    
    return False

# Modified main function
def scrape_company(company, location=None, manual_domain=None):
    global process_log
    process_log = []
    query = f"{company} {location}" if location else company
    custom_print(f"\n🔍 Searching Google for '{query}'...")
    
    domain = manual_domain or find_domain_google(query, company)
    
    if not domain:
        custom_print("❌ No valid domain found matching the company name.")
        return {
            "company_name": company,
            "domain": "N/A",
            "emails": [],
            "social_media": {platform: [] for platform in SOCIAL_REGEXES.keys()},
            "meta_description": "N/A",
            "about_content": "N/A",
            "phone_numbers": [],
            "addresses": [],
            "links": [],
            "country": "N/A",
            "business_nature": "N/A",
            "company_links": [],
            "process_log": process_log
        }
    
    if not domain.startswith(("http://", "https://")):
        domain = f"https://{domain}"
    if not domain.endswith("/"):
        domain += "/"
    
    custom_print(f"✅ Using verified domain: {domain}")
    custom_print(f"📡 Starting full-site scraping from: {domain}\n")

    emails, socials, meta_desc, about_content, phones, addresses, links, location_info, business_nature, footer_contents, debug_info = crawl_website(domain)

    fallback_debug_logs = []
    company_links = []
    # Only perform fallback if emails are missing or some social platforms have no links
    missing_socials = [platform for platform in SOCIAL_REGEXES.keys() if not socials.get(platform)]
    if not emails or missing_socials:
        custom_print(f"\n⚠️ Missing some data (Emails: {not emails}, Missing socials: {missing_socials}). Performing fallback Google searches...")
        
        if not emails:
            email_query = f"{company} {location} email" if location else f"{company} email"
            custom_print(f"🔍 Searching Google for '{email_query}'...")
            fallback_emails, email_company_links, email_debug_log = fallback_google_search(email_query, company, location or "")
            emails.update(fallback_emails)
            company_links.extend(email_company_links)
            custom_print(f"📧 Found {len(fallback_emails)} emails in fallback search: {fallback_emails}")
            custom_print(f"🔗 Found {len(email_company_links)} company links in email search: {email_company_links}")
            fallback_debug_logs.append({"query": email_query, "log": email_debug_log})

        if missing_socials:
            custom_print(f"🔍 Searching Google for social media: {missing_socials}")
            for platform in missing_socials:
                social_query = f"{company} {location} {platform}" if location else f"{company} {platform}"
                custom_print(f"🔍 Searching Google for '{social_query}'...")
                fallback_socials, social_company_links, social_debug_log = fallback_google_search(social_query, company, location or "", [platform])
                if fallback_socials[platform]:
                    socials[platform] = list(set(socials.get(platform, []) + fallback_socials[platform]))
                    custom_print(f"🔗 Found {len(fallback_socials[platform])} {platform} links: {fallback_socials[platform]}")
                else:
                    custom_print(f"🔗 No {platform} links found in fallback search.")
                company_links.extend(social_company_links)
                custom_print(f"🔗 Found {len(social_company_links)} company links in {platform} search: {social_company_links}")
                fallback_debug_logs.append({"query": social_query, "log": social_debug_log})

    company_links = sorted(list(set(company_links)))
    debug_info.append({
        "url": "fallback_searches",
        "status": "fallback",
        "fallback_logs": fallback_debug_logs
    })

    # Convert phone set to list for the results
    phone_list = list(phones)

    results = save_results(
        domain, 
        emails, 
        socials, 
        meta_desc, 
        about_content, 
        phone_list,  # Pass list of phone strings
        addresses, 
        links, 
        location_info,
        business_nature, 
        footer_contents, 
        debug_info, 
        process_log, 
        company_links, 
        company
    )
    return results

# Excel Export function
def generate_excel(results_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Scraper Results"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    headers = [
        'Company Name', 'Domain', 'Country', 'Business Nature',
        'Meta Description', 'Emails', 'Phone Numbers',
        'Facebook', 'Instagram', 'LinkedIn', 'YouTube', 'Pinterest'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row_idx, result in enumerate(results_data, 2):
        social_media = result.get('social_media', {})
        facebook_links = '\n'.join(social_media.get('facebook', []))
        instagram_links = '\n'.join(social_media.get('instagram', []))
        linkedin_links = '\n'.join(social_media.get('linkedin', []))
        youtube_links = '\n'.join(social_media.get('youtube', []))
        pinterest_links = '\n'.join(social_media.get('pinterest', []))
        
        emails = '\n'.join(result.get('emails', []))
        phones = '\n'.join(result.get('phone_numbers', []))
        
        row_data = [
            result.get('company_name', 'N/A'),
            result.get('domain', 'N/A'),
            result.get('country', 'N/A'),
            result.get('business_nature', 'N/A'),
            result.get('meta_description', 'N/A'),
            emails or 'N/A',
            phones or 'N/A',
            facebook_links or 'N/A',
            instagram_links or 'N/A',
            linkedin_links or 'N/A',
            youtube_links or 'N/A',
            pinterest_links or 'N/A'
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = border

    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, len(results_data) + 2):
            cell_value = str(ws[f"{column_letter}{row}"].value or "")
            lines = cell_value.split('\n')
            max_line_length = max(len(line) for line in lines) if lines else 0
            max_length = max(max_length, max_line_length)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    for row in range(2, len(results_data) + 2):
        max_lines = 1
        for col in range(1, len(headers) + 1):
            cell_value = str(ws.cell(row=row, column=col).value or "")
            lines = len(cell_value.split('\n'))
            max_lines = max(max_lines, lines)
        ws.row_dimensions[row].height = max_lines * 15

    ws.row_dimensions[1].height = 30
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

# Flask Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/search', methods=['POST'])
def search():
    global current_results
    try:
        search_type = request.form.get('search_type')
        results = []
        
        if search_type == 'single':
            company = request.form.get('company', '').strip()
            location = request.form.get('location', '').strip()
            manual_domain = request.form.get('manual_domain', '').strip()
            
            if not company:
                return jsonify({
                    'error': 'Please provide company name.',
                    'results': [],
                    'errors': []
                }), 400
            
            custom_print(f"Processing single search for {company} in {location or 'N/A'}")
            result = scrape_company(company, location or None, manual_domain or None)
            results.append(result)
            
        elif search_type in ['batch_5', 'batch_10']:
            batch_size = 5 if search_type == 'batch_5' else 10
            batch_errors = []
            
            for i in range(batch_size):
                company = request.form.get(f'company_{i}', '').strip()
                location = request.form.get(f'location_{i}', '').strip()
                
                if company:
                    custom_print(f"Processing batch search {i+1}/{batch_size} for {company} in {location or 'N/A'}")
                    result = scrape_company(company, location or None)
                    result['serial_no'] = i + 1
                    results.append(result)
                else:
                    custom_print(f"Skipping batch entry {i+1} due to missing company name")
                    batch_errors.append({
                        'batch_index': i + 1,
                        'error': 'Missing company name'
                    })
            
            if not results and batch_errors:
                return jsonify({
                    'error': 'Please provide at least one valid company name for batch search.',
                    'results': [],
                    'errors': batch_errors
                }), 400

        # Aggregate error details from debug_info
        aggregated_errors = []
        for result in results:
            debug_info = result.get('process_log', [])
            for debug_entry in debug_info:
                if 'error_details' in debug_entry and debug_entry['error_details']:
                    for error in debug_entry['error_details']:
                        aggregated_errors.append({
                            'url': debug_entry['url'],
                            'method': error.get('method', 'unknown'),
                            'attempt': error.get('attempt', 'unknown'),
                            'error_type': error.get('error_type', 'unknown'),
                            'message': error.get('message', 'No error message provided'),
                            'http_status': error.get('http_status', 'Unknown'),
                            'current_url': error.get('current_url', 'Unknown')
                        })

        return jsonify({
            'success': True,
            'results': results,
            'batch': search_type.startswith('batch'),
            'errors': aggregated_errors  # Include detailed error information
        })
        
    except Exception as e:
        custom_print(f"Error in search route: {str(e)}")
        return jsonify({
            'error': f'An error occurred during the search: {str(e)}',
            'results': [],
            'errors': [{'error_type': type(e).__name__, 'message': str(e)}]
        }), 500

@app.route('/export/excel', methods=['POST'])
def export_excel():
    try:
        data = request.get_json()
        results_data = data.get('results', []) if data else current_results
        
        if not results_data:
            return jsonify({'error': 'No results to export'}), 400
        
        excel_content = generate_excel(results_data)
        
        response = Response(
            excel_content,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=company_scraper_results_{time.strftime("%Y%m%d_%H%M%S")}.xlsx'
            }
        )
        
        return response
        
    except Exception as e:
        custom_print(f"Error in Excel export: {str(e)}")
        return jsonify({'error': f'Failed to export Excel: {str(e)}'}), 500

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)