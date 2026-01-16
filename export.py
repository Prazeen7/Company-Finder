import io
import json
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import DEBUG_DIR
from utils import custom_print

def save_results(domain, emails, socials, meta_description, about_content, phones, addresses, links, location_info, business_nature, footer_contents, debug_info, process_log, company_links=None, company_name="Unknown"):
    results = {
        "company_name": company_name,
        "domain": domain,
        "emails": list(emails),
        "social_media": socials,
        "meta_description": meta_description or "N/A",
        "about_content": about_content or "N/A",
        "phone_numbers": list(phones),
        "addresses": list(addresses),
        "links": list(links),
        "country": location_info.get("country", "N/A"),  # Only country now
        "business_nature": business_nature or "N/A",
        "company_links": company_links or [],
        "process_log": process_log
    }

    output_file = os.path.join(DEBUG_DIR, f"scrape_results_{company_name.replace(' ', '_')}.json")
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2)
    custom_print(f"📝 Core scraped information and process logs saved to {output_file}")

    with open(os.path.join(DEBUG_DIR, f"footer_content_{company_name.replace(' ', '_')}.json"), "w", encoding="utf-8") as f:
        json.dump([
            {
                "url": fc["url"],
                "html": fc["html"],
                "text": fc["text"]
            } for fc in footer_contents
        ], f, indent=2)
    footer_filename = f"footer_content_{company_name.replace(' ', '_')}.json"
    custom_print(f"📝 Footer content saved to {os.path.join(DEBUG_DIR, footer_filename)}")

    with open(os.path.join(DEBUG_DIR, f"debug_info_{company_name.replace(' ', '_')}.json"), "w", encoding="utf-8") as f:
        json.dump([
            {
                "url": di["url"],
                "status": di.get("status", "unknown"),
                "dynamic_html": di.get("dynamic_html", ""),
                "links": di.get("links", []),
                "scripts": di.get("scripts", []),
                "raw_socials": di.get("raw_socials", []),
                "raw_emails": di.get("raw_emails", []),
                "raw_phones": di.get("raw_phones", []),
                "raw_addresses": di.get("raw_addresses", []),
                "shadow_links": di.get("shadow_links", []),
                "shadow_content": di.get("shadow_content", []),
                "footer_tracking": di.get("footer_tracking", []),
                "meta_tags": di.get("meta_tags", []),
                "page_content": di.get("page_content", {}),
                "all_links": di.get("all_links", []),
                "consolidated_links": di.get("consolidated_links", []),
                "fallback_logs": di.get("fallback_logs", []),
                "alerts": di.get("alerts", []),
                "iframes": di.get("iframes", [])
            } for di in debug_info
        ], f, indent=2)
    debug_filename = f"debug_info_{company_name.replace(' ', '_')}.json"
    custom_print(f"📝 Debug information saved to {os.path.join(DEBUG_DIR, debug_filename)}")

    return results

def generate_excel(results_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Company Scraper Results"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    
    cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    headers = [
        'Company Name', 'Domain', 'Country', 'Business Nature',
        'Meta Description', 'Emails', 'Phone Numbers',
        'Facebook', 'Instagram', 'LinkedIn', 'YouTube', 'Pinterest'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    for row_idx, result in enumerate(results_data, 2):
        social_media = result.get('social_media', {})
        facebook_links = '\n'.join(social_media.get('facebook', []))
        instagram_links = '\n'.join(social_media.get('instagram', []))
        linkedin_links = '\n'.join(social_media.get('linkedin', []))
        youtube_links = '\n'.join(social_media.get('youtube', []))
        pinterest_links = '\n'.join(social_media.get('pinterest', []))
        
        emails = '\n'.join(result.get('emails', []))
        phones = '\n'.join(result.get('phone_numbers', []))
        
        row_data = [
            result.get('company_name', 'N/A'),
            result.get('domain', 'N/A'),
            result.get('country', 'N/A'),
            result.get('business_nature', 'N/A'),
            result.get('meta_description', 'N/A'),
            emails or 'N/A',
            phones or 'N/A',
            facebook_links or 'N/A',
            instagram_links or 'N/A',
            linkedin_links or 'N/A',
            youtube_links or 'N/A',
            pinterest_links or 'N/A'
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = border

    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, len(results_data) + 2):
            cell_value = str(ws[f"{column_letter}{row}"].value or "")
            lines = cell_value.split('\n')
            max_line_length = max(len(line) for line in lines) if lines else 0
            max_length = max(max_length, max_line_length)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    for row in range(2, len(results_data) + 2):
        max_lines = 1
        for col in range(1, len(headers) + 1):
            cell_value = str(ws.cell(row=row, column=col).value or "")
            lines = len(cell_value.split('\n'))
            max_lines = max(max_lines, lines)
        ws.row_dimensions[row].height = max_lines * 15

    ws.row_dimensions[1].height = 30
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()